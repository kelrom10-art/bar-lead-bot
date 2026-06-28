#!/usr/bin/env python3
"""
בוט ניהול לידים לבר 🍸
SQLite (אחסון מקומי) + APScheduler (תזכורות)
"""

import io, os, sqlite3, logging, threading, uuid
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from http.server import HTTPServer, BaseHTTPRequestHandler
from contextlib import contextmanager

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, KeyboardButton,
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler, ConversationHandler,
    CallbackQueryHandler, ContextTypes, filters,
)
from apscheduler.schedulers.asyncio import AsyncIOScheduler

# ── הגדרות ───────────────────────────────────────────────────────────────────
BOT_TOKEN     = os.getenv("BOT_TOKEN", "")
OWNER_CHAT_ID = int(os.getenv("OWNER_CHAT_ID", "0"))
TZ            = ZoneInfo("Asia/Jerusalem")
DB_PATH       = os.path.join(os.path.dirname(__file__), "leads.db")

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# ── ConversationHandler states ───────────────────────────────────────────────
ASK_NAME, ASK_PHONE, ASK_TIME = range(3)

# ── סטטוסים ─────────────────────────────────────────────────────────────────
ST_NEW          = "new"
ST_PRE_REMINDED = "pre_reminded"
ST_CALLED       = "called"
ST_WON          = "won"
ST_LOST         = "lost"
ST_FOLLOW_UP    = "follow_up"
ST_OLD          = "old"

STATUS_EMOJI = {
    ST_NEW: "🆕", ST_PRE_REMINDED: "⏰", ST_CALLED: "📞",
    ST_WON: "✅", ST_LOST: "❌", ST_FOLLOW_UP: "🔄", ST_OLD: "📁",
}

# ── תפריט ראשי ───────────────────────────────────────────────────────────────
MENU_BUTTONS = {
    "➕ הוספת ליד", "📋 לידים ישנים", "📊 סיכום", "⏰ תזכורות פעילות",
    "📅 היום", "🔍 חיפוש ליד", "📤 יצוא Excel", "✏️ עריכת ליד",
}

MAIN_MENU = ReplyKeyboardMarkup(
    [
        [KeyboardButton("➕ הוספת ליד"),      KeyboardButton("📋 לידים ישנים")],
        [KeyboardButton("📊 סיכום"),           KeyboardButton("⏰ תזכורות פעילות")],
        [KeyboardButton("📅 היום"),            KeyboardButton("🔍 חיפוש ליד")],
        [KeyboardButton("📤 יצוא Excel"),      KeyboardButton("✏️ עריכת ליד")],
    ],
    resize_keyboard=True,
    is_persistent=True,
)

# ═══════════════════════════════════════════════
#  SQLite helpers
# ═══════════════════════════════════════════════

@contextmanager
def _db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()

def init_db():
    try:
        with _db() as conn:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS leads (
                    id             TEXT PRIMARY KEY,
                    name           TEXT NOT NULL,
                    phone          TEXT NOT NULL,
                    call_time      TEXT,
                    status         TEXT DEFAULT 'new',
                    sale_amount    REAL DEFAULT 0,
                    notes          TEXT DEFAULT '',
                    created_at     TEXT,
                    updated_at     TEXT,
                    follow_up_time TEXT
                );
            """)
        logger.info("DB ✓")
    except sqlite3.OperationalError as e:
        logger.warning(f"DB פגום, יוצר מחדש: {e}")
        for ext in ("", "-journal", "-wal", "-shm"):
            p = DB_PATH + ext
            try:
                os.remove(p)
            except OSError:
                pass
        with _db() as conn:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS leads (
                    id             TEXT PRIMARY KEY,
                    name           TEXT NOT NULL,
                    phone          TEXT NOT NULL,
                    call_time      TEXT,
                    status         TEXT DEFAULT 'new',
                    sale_amount    REAL DEFAULT 0,
                    notes          TEXT DEFAULT '',
                    created_at     TEXT,
                    updated_at     TEXT,
                    follow_up_time TEXT
                );
            """)
        logger.info("DB נוצר מחדש ✓")

# ═══════════════════════════════════════════════
#  CRUD
# ═══════════════════════════════════════════════

def _now():
    return datetime.now(TZ).strftime("%Y-%m-%dT%H:%M:%S")

def add_lead(name, phone, call_dt):
    lid = str(uuid.uuid4())[:8].upper()
    with _db() as conn:
        conn.execute(
            "INSERT INTO leads (id, name, phone, call_time, status, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (lid, name, phone, call_dt.strftime("%Y-%m-%dT%H:%M:%S"), ST_NEW, _now(), _now())
        )
    return lid

def update_lead(lid, **kw):
    kw["updated_at"] = _now()
    cols = ", ".join(f"{k}=?" for k in kw)
    vals = list(kw.values()) + [lid]
    with _db() as conn:
        conn.execute(f"UPDATE leads SET {cols} WHERE id=?", vals)

def delete_lead(lid):
    with _db() as conn:
        conn.execute("DELETE FROM leads WHERE id=?", (lid,))

def get_lead(lid):
    with _db() as conn:
        row = conn.execute("SELECT * FROM leads WHERE id=?", (lid,)).fetchone()
    return dict(row) if row else None

def _all():
    with _db() as conn:
        rows = conn.execute("SELECT * FROM leads ORDER BY call_time").fetchall()
    return [dict(r) for r in rows]

def get_active_leads():
    with _db() as conn:
        rows = conn.execute(
            "SELECT * FROM leads WHERE status IN (?,?,?,?) ORDER BY call_time",
            (ST_NEW, ST_PRE_REMINDED, ST_CALLED, ST_FOLLOW_UP)
        ).fetchall()
    return [dict(r) for r in rows]

def get_today_leads():
    today = datetime.now(TZ).strftime("%Y-%m-%d")
    with _db() as conn:
        rows = conn.execute(
            "SELECT * FROM leads WHERE call_time LIKE ? AND status IN (?,?,?,?) ORDER BY call_time",
            (f"{today}%", ST_NEW, ST_PRE_REMINDED, ST_CALLED, ST_FOLLOW_UP)
        ).fetchall()
    return [dict(r) for r in rows]

def get_old_leads():
    with _db() as conn:
        rows = conn.execute(
            "SELECT * FROM leads WHERE status=? ORDER BY updated_at DESC", (ST_OLD,)
        ).fetchall()
    return [dict(r) for r in rows]

def search_leads(query):
    q = f"%{query.lower()}%"
    with _db() as conn:
        rows = conn.execute(
            "SELECT * FROM leads WHERE LOWER(name) LIKE ? OR phone LIKE ?", (q, q)
        ).fetchall()
    return [dict(r) for r in rows]

def get_leads_pre_reminder():
    now = datetime.now(TZ)
    out = []
    with _db() as conn:
        rows = conn.execute(
            "SELECT * FROM leads WHERE status=? AND call_time IS NOT NULL", (ST_NEW,)
        ).fetchall()
    for r in rows:
        try:
            dt   = datetime.strptime(r["call_time"][:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=TZ)
            diff = (dt - now).total_seconds() / 60
            if 0 <= diff <= 20:
                out.append(dict(r))
        except Exception:
            pass
    return out

def get_leads_post_call():
    now = datetime.now(TZ)
    out = []
    with _db() as conn:
        rows = conn.execute(
            "SELECT * FROM leads WHERE status IN (?,?) AND call_time IS NOT NULL",
            (ST_NEW, ST_PRE_REMINDED)
        ).fetchall()
    for r in rows:
        try:
            dt   = datetime.strptime(r["call_time"][:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=TZ)
            diff = (now - dt).total_seconds() / 60
            if diff >= 30:
                out.append(dict(r))
        except Exception:
            pass
    return out

def get_leads_followup_due():
    now = _now()
    with _db() as conn:
        rows = conn.execute(
            "SELECT * FROM leads WHERE status=? AND follow_up_time IS NOT NULL AND follow_up_time <= ?",
            (ST_FOLLOW_UP, now)
        ).fetchall()
    return [dict(r) for r in rows]

def get_stats(since_dt):
    since = since_dt.strftime("%Y-%m-%dT%H:%M:%S")
    with _db() as conn:
        all_r  = conn.execute("SELECT * FROM leads WHERE created_at >= ?", (since,)).fetchall()
        won    = conn.execute("SELECT * FROM leads WHERE status=? AND created_at >= ?", (ST_WON, since)).fetchall()
        lost   = conn.execute("SELECT COUNT(*) FROM leads WHERE status=? AND created_at >= ?", (ST_LOST, since)).fetchone()[0]
        active = conn.execute(
            "SELECT COUNT(*) FROM leads WHERE status IN (?,?,?,?) AND created_at >= ?",
            (ST_NEW, ST_PRE_REMINDED, ST_CALLED, ST_FOLLOW_UP, since)
        ).fetchone()[0]
        old    = conn.execute("SELECT COUNT(*) FROM leads WHERE status=? AND created_at >= ?", (ST_OLD, since)).fetchone()[0]
    return {
        "total":      len(all_r),
        "won_count":  len(won),
        "won_amount": sum(r["sale_amount"] or 0 for r in won),
        "lost":       lost,
        "active":     active,
        "old":        old,
    }

# ═══════════════════════════════════════════════
#  Excel export
# ═══════════════════════════════════════════════

def build_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "לידים"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2B579A")
    center      = Alignment(horizontal="center")

    col_labels = ["מזהה", "שם", "טלפון", "שעת שיחה", "סטטוס", "סכום מכירה", "הערות", "נוצר"]
    col_keys   = ["id",   "name", "phone", "call_time", "status", "sale_amount", "notes", "created_at"]
    status_heb = {
        ST_NEW: "חדש", ST_PRE_REMINDED: "תזכורת נשלחה", ST_CALLED: "עובד",
        ST_WON: "נסגר ✅", ST_LOST: "אבד ❌", ST_FOLLOW_UP: "המשך טיפול", ST_OLD: "ישן",
    }

    for ci, label in enumerate(col_labels, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center

    for ri, row in enumerate(_all(), 2):
        for ci, key in enumerate(col_keys, 1):
            val = row.get(key, "")
            if key == "status":
                val = status_heb.get(str(val), str(val))
            elif key in ("call_time", "created_at") and val:
                try:
                    val = datetime.strptime(str(val)[:19], "%Y-%m-%dT%H:%M:%S").strftime("%d/%m/%Y %H:%M")
                except Exception:
                    pass
            elif key == "sale_amount" and val:
                try:
                    val = f"₪{float(val):,.0f}"
                except Exception:
                    pass
            ws.cell(row=ri, column=ci, value=str(val) if val else "")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ═══════════════════════════════════════════════
#  Parse time string
# ═══════════════════════════════════════════════

def parse_time(text):
    now = datetime.now(TZ)
    t   = text.strip()
    if "מחר" in t:
        base = (now + timedelta(days=1)).replace(second=0, microsecond=0)
        for w in t.split():
            if ":" in w:
                try:
                    h, m = map(int, w.split(":"))
                    return base.replace(hour=h, minute=m)
                except Exception:
                    pass
        return base.replace(hour=9, minute=0)
    if "שעות" in t or "שעה" in t:
        for w in t.split():
            if w.isdigit():
                return now + timedelta(hours=int(w))
        return now + timedelta(hours=1)
    if "דקות" in t or "דקה" in t:
        for w in t.split():
            if w.isdigit():
                return now + timedelta(minutes=int(w))
        return now + timedelta(minutes=30)
    parts = t.split()
    if len(parts) == 2 and "/" in parts[0] and ":" in parts[1]:
        try:
            day, month = map(int, parts[0].split("/"))
            h, m       = map(int, parts[1].split(":"))
            dt = now.replace(month=month, day=day, hour=h, minute=m, second=0, microsecond=0)
            if dt < now:
                dt = dt.replace(year=now.year + 1)
            return dt
        except Exception:
            pass
    if ":" in t:
        try:
            h, m = map(int, t.split(":"))
            dt   = now.replace(hour=h, minute=m, second=0, microsecond=0)
            return dt if dt > now else dt + timedelta(days=1)
        except Exception:
            pass
    return None

def validate_phone(phone):
    digits = "".join(c for c in phone if c.isdigit())
    if digits.startswith("972"):
        digits = "0" + digits[3:]
    return len(digits) == 10 and digits.startswith("05")

# ═══════════════════════════════════════════════
#  Inline keyboards
# ═══════════════════════════════════════════════

def outcome_kb(lid):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ נסגר! 🎉",       callback_data=f"out_{lid}_won")],
        [InlineKeyboardButton("⏰ נדחה — קבע זמן", callback_data=f"out_{lid}_snooze"),
         InlineKeyboardButton("🔄 ליד ישן",        callback_data=f"out_{lid}_old")],
        [InlineKeyboardButton("❌ לא רלוונטי",     callback_data=f"out_{lid}_lost")],
    ])

def snooze_kb(lid):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("בעוד שעה ⏰",  callback_data=f"snz_{lid}_1h"),
         InlineKeyboardButton("מחר 9:00 🌅",  callback_data=f"snz_{lid}_tomorrow")],
        [InlineKeyboardButton("קבע ידנית ✏️", callback_data=f"snz_{lid}_manual")],
    ])

def edit_field_kb(lid):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👤 שינוי שם",       callback_data=f"edf_{lid}_name")],
        [InlineKeyboardButton("📱 שינוי טלפון",     callback_data=f"edf_{lid}_phone")],
        [InlineKeyboardButton("📅 שינוי שעת שיחה", callback_data=f"edf_{lid}_time")],
        [InlineKeyboardButton("🗑 מחיקת ליד",      callback_data=f"edf_{lid}_delete")],
    ])

# ═══════════════════════════════════════════════
#  ConversationHandler — הוספת ליד
# ═══════════════════════════════════════════════

async def add_start(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await u.message.reply_text(
        "👤 *מה שם הלקוח?*\n\n_שלח /cancel לביטול_",
        parse_mode="Markdown")
    return ASK_NAME

async def got_name(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = u.message.text.strip()
    if name in MENU_BUTTONS:
        return await cancel_add(u, ctx)
    ctx.user_data["name"] = name
    await u.message.reply_text(
        f"✅ שם: *{name}*\n\n📱 *מה מספר הטלפון?*\n_לדוגמא: 0501234567_",
        parse_mode="Markdown")
    return ASK_PHONE

async def got_phone(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    phone = u.message.text.strip()
    if phone in MENU_BUTTONS:
        return await cancel_add(u, ctx)
    if not validate_phone(phone):
        await u.message.reply_text(
            "❌ מספר טלפון לא תקין!\nנא להכניס מספר ישראלי: *0501234567*",
            parse_mode="Markdown")
        return ASK_PHONE
    ctx.user_data["phone"] = phone
    await u.message.reply_text(
        f"✅ טלפון: *{phone}*\n\n"
        "📅 *מתי השיחה?*\n"
        "_לדוגמא:_\n"
        "`16:30` — היום בשעה 16:30\n"
        "`מחר 10:00` — מחר בעשר\n"
        "`25/06 15:30` — תאריך ספציפי\n"
        "`בעוד שעה` — עוד שעה מעכשיו",
        parse_mode="Markdown")
    return ASK_TIME

async def got_time(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = u.message.text.strip()
    if text in MENU_BUTTONS:
        return await cancel_add(u, ctx)
    call_dt = parse_time(text)
    if not call_dt:
        await u.message.reply_text(
            "❓ לא הבנתי. נסה:\n`16:30` | `מחר 10:00` | `25/06 15:30`",
            parse_mode="Markdown")
        return ASK_TIME
    name  = ctx.user_data["name"]
    phone = ctx.user_data["phone"]
    lid   = add_lead(name, phone, call_dt)
    ctx.user_data.clear()
    await u.message.reply_text(
        f"🎯 *ליד נוסף בהצלחה!*\n\n"
        f"👤 *{name}*\n"
        f"📱 {phone}\n"
        f"📅 שיחה: *{call_dt.strftime('%d/%m/%Y בשעה %H:%M')}*\n\n"
        f"⏰ תקבל תזכורת 20 דקות לפני",
        parse_mode="Markdown",
        reply_markup=MAIN_MENU,
    )
    return ConversationHandler.END

async def cancel_add(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    if u.message:
        await u.message.reply_text("❌ הוספת ליד בוטלה.", reply_markup=MAIN_MENU)
    return ConversationHandler.END

# ═══════════════════════════════════════════════
#  /start
# ═══════════════════════════════════════════════

async def cmd_start(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await u.message.reply_text(
        "👋 *שלום! בוט ניהול הלידים שלך* 🍸\n\nבחר פעולה:",
        parse_mode="Markdown",
        reply_markup=MAIN_MENU,
    )

# ═══════════════════════════════════════════════
#  Menu handlers
# ═══════════════════════════════════════════════

async def show_today(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    leads = get_today_leads()
    if not leads:
        await u.message.reply_text("📅 אין שיחות מתוכננות להיום! 🎉", reply_markup=MAIN_MENU)
        return
    lines = [f"📅 *שיחות להיום — {datetime.now(TZ).strftime('%d/%m/%Y')} ({len(leads)}):*\n"]
    for ld in leads:
        ct = str(ld.get("call_time", ""))
        t  = ""
        try:
            t = datetime.strptime(ct[:19], "%Y-%m-%dT%H:%M:%S").strftime("%H:%M")
        except Exception:
            pass
        lines.append(
            f"🕐 *{t}* — {STATUS_EMOJI.get(ld.get('status'), '❓')} "
            f"*{ld['name']}* | 📱{ld.get('phone', '—')}"
        )
    await u.message.reply_text("\n".join(lines), parse_mode="Markdown", reply_markup=MAIN_MENU)

async def show_old_leads(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    old = get_old_leads()
    if not old:
        await u.message.reply_text("📁 אין לידים ישנים.", reply_markup=MAIN_MENU)
        return
    await u.message.reply_text(f"📁 *לידים ישנים ({len(old)}):*", parse_mode="Markdown")
    for ld in old:
        ct       = str(ld.get("call_time", "") or "")
        date_str = ct[:10].replace("-", "/") if ct else "—"
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📞 לחזור אליו", callback_data=f"recall_{ld['id']}"),
             InlineKeyboardButton("✅ נסגר",        callback_data=f"out_{ld['id']}_won")],
            [InlineKeyboardButton("❌ לא רלוונטי", callback_data=f"out_{ld['id']}_lost")],
        ])
        await u.message.reply_text(
            f"📁 *{ld['name']}*\n📱 {ld.get('phone', '—')}\n📅 שיחה אחרונה: {date_str}",
            reply_markup=kb, parse_mode="Markdown")

async def show_reminders(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    active = get_active_leads()
    if not active:
        await u.message.reply_text("✅ אין תזכורות פעילות!", reply_markup=MAIN_MENU)
        return
    lines = [f"⏰ *תזכורות פעילות ({len(active)}):*\n"]
    for ld in active:
        ct = str(ld.get("call_time") or ld.get("follow_up_time") or "")
        t  = ""
        if ct:
            try:
                t = f" | ⏰ {datetime.strptime(ct[:19], '%Y-%m-%dT%H:%M:%S').strftime('%d/%m %H:%M')}"
            except Exception:
                pass
        lines.append(
            f"{STATUS_EMOJI.get(ld.get('status'), '❓')} *{ld['name']}* "
            f"— 📱{ld.get('phone', '—')}{t}"
        )
    await u.message.reply_text("\n".join(lines), parse_mode="Markdown", reply_markup=MAIN_MENU)

async def show_summary(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    now     = datetime.now(TZ)
    weekly  = get_stats(now - timedelta(days=7))
    monthly = get_stats(now.replace(day=1, hour=0, minute=0, second=0))
    rate    = lambda s: f"{s['won_count']/s['total']*100:.0f}%" if s["total"] else "—"
    await u.message.reply_text(
        "📊 *סיכום ביצועים*\n\n"
        f"📅 *שבוע אחרון:*\n"
        f"  לידים: {weekly['total']} | ✅ נסגרו: {weekly['won_count']} | ❌ אבדו: {weekly['lost']}\n"
        f"  💰 הכנסות: ₪{weekly['won_amount']:,.0f} | המרה: {rate(weekly)}\n\n"
        f"📆 *החודש הנוכחי:*\n"
        f"  לידים: {monthly['total']} | ✅ נסגרו: {monthly['won_count']} | ❌ אבדו: {monthly['lost']}\n"
        f"  💰 הכנסות: ₪{monthly['won_amount']:,.0f} | המרה: {rate(monthly)}\n\n"
        f"🟢 פעילים: {len(get_active_leads())} | 📁 ישנים: {len(get_old_leads())}",
        parse_mode="Markdown",
        reply_markup=MAIN_MENU,
    )

async def start_search(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["state"] = "searching"
    await u.message.reply_text("🔍 *חיפוש ליד*\n\nהכנס שם או מספר טלפון:", parse_mode="Markdown")

async def export_excel(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await u.message.reply_text("⏳ מכין קובץ Excel...", reply_markup=MAIN_MENU)
    try:
        buf   = build_excel()
        fname = f"leads_{datetime.now(TZ).strftime('%d%m%Y_%H%M')}.xlsx"
        await u.message.reply_document(
            document=buf,
            filename=fname,
            caption=f"📊 כל הלידים — {datetime.now(TZ).strftime('%d/%m/%Y %H:%M')}")
    except Exception as e:
        logger.error(f"export_excel: {e}")
        await u.message.reply_text(f"❌ שגיאה ביצירת קובץ: {e}")

async def start_edit(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    active = get_active_leads()
    if not active:
        await u.message.reply_text("❌ אין לידים פעילים לעריכה.", reply_markup=MAIN_MENU)
        return
    buttons = []
    for ld in active[:10]:
        ct = str(ld.get("call_time", ""))
        t  = ""
        try:
            t = f" {datetime.strptime(ct[:19], '%Y-%m-%dT%H:%M:%S').strftime('%d/%m %H:%M')}"
        except Exception:
            pass
        buttons.append([InlineKeyboardButton(
            f"{STATUS_EMOJI.get(ld.get('status'), '❓')} {ld['name']}{t}",
            callback_data=f"edit_{ld['id']}"
        )])
    await u.message.reply_text(
        "✏️ *בחר ליד לעריכה:*",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="Markdown")

# ═══════════════════════════════════════════════
#  Main text router
# ═══════════════════════════════════════════════

async def handle_text(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text  = u.message.text.strip()
    state = ctx.user_data.get("state")

    if state == "ask_amount":
        lid = ctx.user_data.get("lid")
        try:
            amount = float(text.replace("₪", "").replace(",", ""))
            ld     = get_lead(lid)
            update_lead(lid, status=ST_WON, sale_amount=amount)
            ctx.user_data.clear()
            await u.message.reply_text(
                f"🎉 *כל הכבוד!*\n*{ld['name']}* — עסקה נסגרה\n💰 *₪{amount:,.0f}*",
                parse_mode="Markdown", reply_markup=MAIN_MENU)
        except ValueError:
            await u.message.reply_text("❓ הכנס סכום בשקלים (לדוגמא: `3500`)", parse_mode="Markdown")
        return

    if state in ("ask_snooze", "ask_recall"):
        lid = ctx.user_data.get("lid")
        dt  = parse_time(text)
        if dt:
            if state == "ask_recall":
                update_lead(lid, status=ST_NEW,
                            call_time=dt.strftime("%Y-%m-%dT%H:%M:%S"),
                            follow_up_time=None)
            else:
                update_lead(lid, status=ST_FOLLOW_UP,
                            follow_up_time=dt.strftime("%Y-%m-%dT%H:%M:%S"))
            ctx.user_data.clear()
            await u.message.reply_text(
                f"⏰ תזכורת נקבעה ל-*{dt.strftime('%d/%m/%Y %H:%M')}* ✓",
                parse_mode="Markdown", reply_markup=MAIN_MENU)
        else:
            await u.message.reply_text(
                "❓ נסה: `16:30` | `מחר 10:00` | `25/06 15:00`", parse_mode="Markdown")
        return

    if state == "searching":
        results = search_leads(text)
        ctx.user_data.clear()
        if not results:
            await u.message.reply_text(
                f"🔍 לא נמצאו לידים עבור: *{text}*",
                parse_mode="Markdown", reply_markup=MAIN_MENU)
            return
        await u.message.reply_text(
            f"🔍 *{len(results)} תוצאות עבור \"{text}\":*", parse_mode="Markdown")
        for ld in results[:8]:
            ct = str(ld.get("call_time", ""))
            t  = ""
            try:
                t = f"\n📅 {datetime.strptime(ct[:19], '%Y-%m-%dT%H:%M:%S').strftime('%d/%m/%Y %H:%M')}"
            except Exception:
                pass
            kb = InlineKeyboardMarkup([[
                InlineKeyboardButton("✏️ עריכה",      callback_data=f"edit_{ld['id']}"),
                InlineKeyboardButton("📋 עדכן סטטוס", callback_data=f"out_{ld['id']}_snooze"),
            ]])
            await u.message.reply_text(
                f"{STATUS_EMOJI.get(ld.get('status'), '❓')} *{ld['name']}*\n"
                f"📱 {ld.get('phone', '—')}{t}",
                parse_mode="Markdown", reply_markup=kb)
        return

    if state == "edit_value":
        lid   = ctx.user_data.get("lid")
        field = ctx.user_data.get("field")
        if field == "name":
            update_lead(lid, name=text)
            ctx.user_data.clear()
            await u.message.reply_text(
                f"✅ שם עודכן ל-*{text}*", parse_mode="Markdown", reply_markup=MAIN_MENU)
        elif field == "phone":
            if not validate_phone(text):
                await u.message.reply_text(
                    "❌ מספר לא תקין! נסה שוב (לדוגמא: 0501234567)", parse_mode="Markdown")
                return
            update_lead(lid, phone=text)
            ctx.user_data.clear()
            await u.message.reply_text(
                f"✅ טלפון עודכן ל-*{text}*", parse_mode="Markdown", reply_markup=MAIN_MENU)
        elif field == "time":
            dt = parse_time(text)
            if not dt:
                await u.message.reply_text(
                    "❓ נסה: `16:30` | `מחר 10:00` | `25/06 15:00`", parse_mode="Markdown")
                return
            update_lead(lid, call_time=dt.strftime("%Y-%m-%dT%H:%M:%S"))
            ctx.user_data.clear()
            await u.message.reply_text(
                f"✅ שעת שיחה עודכנה ל-*{dt.strftime('%d/%m/%Y %H:%M')}*",
                parse_mode="Markdown", reply_markup=MAIN_MENU)
        return

    # ── Main menu routing ─────────────────────
    ctx.user_data.clear()
    if "לידים ישנים" in text:
        await show_old_leads(u, ctx)
    elif "סיכום" in text:
        await show_summary(u, ctx)
    elif "תזכורות" in text:
        await show_reminders(u, ctx)
    elif "היום" in text:
        await show_today(u, ctx)
    elif "חיפוש" in text:
        await start_search(u, ctx)
    elif "Excel" in text or "יצוא" in text:
        await export_excel(u, ctx)
    elif "עריכת ליד" in text:
        await start_edit(u, ctx)
    else:
        await u.message.reply_text(
            "💬 לחץ על אחת האפשרויות בתפריט 👇", reply_markup=MAIN_MENU)

# ═══════════════════════════════════════════════
#  Callback handler
# ═══════════════════════════════════════════════

async def callback_handler(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = u.callback_query
    await q.answer()
    data = q.data

    if data.startswith("edit_"):
        lid = data.replace("edit_", "")
        ld  = get_lead(lid)
        if not ld:
            await q.edit_message_text("❌ ליד לא נמצא"); return
        ct = str(ld.get("call_time", ""))
        t  = ""
        try:
            t = datetime.strptime(ct[:19], "%Y-%m-%dT%H:%M:%S").strftime("%d/%m/%Y %H:%M")
        except Exception:
            pass
        await q.edit_message_text(
            f"✏️ *עריכת ליד*\n\n"
            f"👤 {ld['name']}\n📱 {ld.get('phone', '—')}\n📅 {t}\n\n*מה לשנות?*",
            reply_markup=edit_field_kb(lid),
            parse_mode="Markdown")
        return

    if data.startswith("edf_"):
        _, lid, field = data.split("_", 2)
        ld = get_lead(lid)
        if not ld:
            await q.edit_message_text("❌ ליד לא נמצא"); return
        if field == "delete":
            delete_lead(lid)
            await q.edit_message_text(f"🗑 *{ld['name']}* נמחק.", parse_mode="Markdown")
            return
        prompts = {
            "name":  f"👤 הכנס שם חדש עבור *{ld['name']}*:",
            "phone": f"📱 הכנס טלפון חדש עבור *{ld['name']}*:",
            "time":  f"📅 הכנס שעת שיחה חדשה עבור *{ld['name']}*:\n`16:30` | `מחר 10:00` | `25/06 15:00`",
        }
        ctx.user_data.update({"state": "edit_value", "lid": lid, "field": field})
        await q.edit_message_text(prompts.get(field, "?"), parse_mode="Markdown")
        return

    if data.startswith("out_"):
        _, lid, action = data.split("_", 2)
        ld = get_lead(lid)
        if not ld:
            await q.edit_message_text("❌ ליד לא נמצא"); return
        if action == "won":
            ctx.user_data.update({"state": "ask_amount", "lid": lid})
            await q.edit_message_text(
                f"🎉 *כמה שילם {ld['name']}?*\n_(הכנס סכום בשקלים)_",
                parse_mode="Markdown")
        elif action == "lost":
            update_lead(lid, status=ST_LOST)
            await q.edit_message_text(
                f"❌ *{ld['name']}* סומן כלא רלוונטי.", parse_mode="Markdown")
        elif action == "snooze":
            await q.edit_message_text(
                f"⏰ *{ld['name']}* — מתי לחזור אליו?",
                reply_markup=snooze_kb(lid), parse_mode="Markdown")
        elif action == "old":
            update_lead(lid, status=ST_OLD)
            await q.edit_message_text(
                f"📁 *{ld['name']}* הועבר ללידים ישנים.", parse_mode="Markdown")
        return

    if data.startswith("snz_"):
        parts  = data.split("_")
        lid    = parts[1]
        option = "_".join(parts[2:])
        now    = datetime.now(TZ)
        if option == "1h":
            new_dt = now + timedelta(hours=1)
        elif option == "tomorrow":
            new_dt = (now + timedelta(days=1)).replace(hour=9, minute=0, second=0, microsecond=0)
        elif option == "manual":
            ctx.user_data.update({"state": "ask_snooze", "lid": lid})
            ld = get_lead(lid)
            await q.edit_message_text(
                f"⏰ מתי לחזור אל *{ld['name']}*?\n`16:30` | `מחר 10:00` | `בעוד שעה`",
                parse_mode="Markdown")
            return
        else:
            return
        update_lead(lid, status=ST_FOLLOW_UP,
                    follow_up_time=new_dt.strftime("%Y-%m-%dT%H:%M:%S"))
        ld = get_lead(lid)
        await q.edit_message_text(
            f"⏰ *{ld['name']}* — אחזור ב-*{new_dt.strftime('%d/%m %H:%M')}* 👍",
            parse_mode="Markdown")
        return

    if data.startswith("recall_"):
        lid = data.replace("recall_", "")
        ld  = get_lead(lid)
        ctx.user_data.update({"state": "ask_recall", "lid": lid})
        await q.edit_message_text(
            f"📞 מתי השיחה עם *{ld['name']}*?\n`16:30` | `מחר 10:00` | `25/06 15:00`",
            parse_mode="Markdown")


# Scheduled jobs

async def job_check_reminders(app):
    for ld in get_leads_pre_reminder():
        try:
            dt = datetime.strptime(ld["call_time"][:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=TZ)
            await app.bot.send_message(
                chat_id=OWNER_CHAT_ID,
                text=f"⏰ *שיחה בעוד ~20 דקות!*\n\n"
                     f"👤 *{ld['name']}*\n"
                     f"📱 {ld.get('phone', '—')}\n"
                     f"🕐 שיחה בשעה *{dt.strftime('%H:%M')}*",
                parse_mode="Markdown")
            update_lead(ld["id"], status=ST_PRE_REMINDED)
        except Exception as e:
            logger.error(f"pre_reminder [{ld.get('id')}]: {e}")

    for ld in get_leads_post_call():
        try:
            await app.bot.send_message(
                chat_id=OWNER_CHAT_ID,
                text=f"📞 *מה קרה עם {ld['name']}?*\n📱 {ld.get('phone', '—')}",
                reply_markup=outcome_kb(ld["id"]),
                parse_mode="Markdown")
            update_lead(ld["id"], status=ST_CALLED)
        except Exception as e:
            logger.error(f"post_call [{ld.get('id')}]: {e}")

    for ld in get_leads_followup_due():
        try:
            await app.bot.send_message(
                chat_id=OWNER_CHAT_ID,
                text=f"🔄 *זמן לחזור אל {ld['name']}!*\n📱 {ld.get('phone', '—')}",
                reply_markup=outcome_kb(ld["id"]),
                parse_mode="Markdown")
            update_lead(ld["id"], status=ST_CALLED)
        except Exception as e:
            logger.error(f"followup_due [{ld.get('id')}]: {e}")

async def job_morning_briefing(app):
    active = get_active_leads()
    old    = get_old_leads()
    if not active and not old:
        return
    now   = datetime.now(TZ)
    lines = [f"☀️ *בוקר טוב! {now.strftime('%d/%m/%Y')}*\n"]
    if active:
        lines.append(f"📋 *שיחות ותזכורות ({len(active)}):*")
        for ld in active[:10]:
            ct = str(ld.get("call_time") or ld.get("follow_up_time") or "")
            t  = ""
            if ct:
                try:
                    t = f" ⏰{datetime.strptime(ct[:19], '%Y-%m-%dT%H:%M:%S').strftime('%H:%M')}"
                except Exception:
                    pass
            lines.append(f"{STATUS_EMOJI.get(ld.get('status'), '❓')} *{ld['name']}*{t}")
    if old:
        lines.append(f"\n📁 *{len(old)} לידים ישנים*")
    await app.bot.send_message(chat_id=OWNER_CHAT_ID, text="\n".join(lines), parse_mode="Markdown")

async def job_weekly_report(app):
    s    = get_stats(datetime.now(TZ) - timedelta(days=7))
    rate = f"{s['won_count']/s['total']*100:.0f}%" if s["total"] else "—"
    await app.bot.send_message(
        chat_id=OWNER_CHAT_ID,
        text=f"📊 *דוח שבועי*\n\n"
             f"לידים: {s['total']} | ✅ נסגרו: {s['won_count']} | ❌ אבדו: {s['lost']}\n"
             f"💰 הכנסות: ₪{s['won_amount']:,.0f} | המרה: {rate}\n\nשבוע מוצלח! 🍸",
        parse_mode="Markdown")

async def job_monthly_report(app):
    now  = datetime.now(TZ)
    s    = get_stats(now.replace(day=1, hour=0, minute=0, second=0))
    rate = f"{s['won_count']/s['total']*100:.0f}%" if s["total"] else "—"
    await app.bot.send_message(
        chat_id=OWNER_CHAT_ID,
        text=f"📆 *דוח חודשי — {now.strftime('%m/%Y')}*\n\n"
             f"לידים: {s['total']} | ✅ נסגרו: {s['won_count']} | ❌ אבדו: {s['lost']}\n"
             f"💰 הכנסות: ₪{s['won_amount']:,.0f} | המרה: {rate}\n\nחודש מוצלח! 💪",
        parse_mode="Markdown")


def main():
    if not BOT_TOKEN:
        raise ValueError("BOT_TOKEN לא הוגדר!")
    if not OWNER_CHAT_ID:
        raise ValueError("OWNER_CHAT_ID לא הוגדר!")

    init_db()

    async def post_init(application: Application) -> None:
        s = AsyncIOScheduler(timezone="Asia/Jerusalem")
        s.add_job(job_check_reminders,  "interval", minutes=1,                  args=[application])
        s.add_job(job_morning_briefing, "cron",     hour=9,  minute=0,          args=[application])
        s.add_job(job_weekly_report,    "cron",     day_of_week="sun", hour=10, args=[application])
        s.add_job(job_monthly_report,   "cron",     day=1,   hour=10,           args=[application])
        s.start()
        logger.info("Scheduler ✓")

    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()

    menu_regex = (
        "^(📋 לידים ישנים"
        "|📊 סיכום"
        "|⏰ תזכורות פעילות"
        "|📅 היום"
        "|🔍 חיפוש ליד"
        "|📤 יצוא Excel"
        "|✏️ עריכת ליד)$"
    )

    conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^➕ הוספת ליד$"), add_start)],
        states={
            ASK_NAME:  [MessageHandler(filters.TEXT & ~filters.COMMAND, got_name)],
            ASK_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_phone)],
            ASK_TIME:  [MessageHandler(filters.TEXT & ~filters.COMMAND, got_time)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel_add),
            MessageHandler(filters.Regex(menu_regex), cancel_add),
        ],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(conv)
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    logger.info("בוט עולה 🚀")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
