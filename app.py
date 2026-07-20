#!/usr/bin/env python3
"""
Bar Lead Manager v4.0 — Unified Render Deployment
FastAPI (REST + PWA) + Telegram Bot + APScheduler + Multi-User Auth
PostgreSQL via psycopg2

Required environment variables:
  DATABASE_URL      — PostgreSQL connection string
  BOT_TOKEN         — Telegram bot token
  OWNER_CHAT_ID     — Telegram chat ID of the owner
  PORT              — Port to listen on (Render sets this automatically)

Optional environment variables:
  SMTP_HOST         — SMTP server host (e.g. smtp.gmail.com)
  SMTP_PORT         — SMTP port (default 587)
  SMTP_USER         — SMTP email address / username
  SMTP_PASS         — SMTP password or app password
  APP_URL           — Public app URL (e.g. https://bar-lead-bot.onrender.com)
  OPEN_REGISTRATION — Allow public self-registration (default: true)
"""

import asyncio, base64, csv, hashlib, hmac as _hmac, io, json, logging, os
import secrets, smtplib, threading, time, uuid
from contextlib import asynccontextmanager, contextmanager
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from typing import Optional
from zoneinfo import ZoneInfo
import urllib.request as _ureq

try:
    from pywebpush import webpush as _webpush, WebPushException
    from cryptography.hazmat.primitives.asymmetric import ec
    from cryptography.hazmat.primitives import serialization
    WEBPUSH_AVAILABLE = True
except ImportError:
    WEBPUSH_AVAILABLE = False

import psycopg2
import psycopg2.extras
import psycopg2.pool

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse, FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, Field, field_validator
import uvicorn
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, KeyboardButton,
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler, ConversationHandler,
    CallbackQueryHandler, ContextTypes, filters,
)
from apscheduler.schedulers.asyncio import AsyncIOScheduler

# ── Config ────────────────────────────────────────────────────────────────────
BOT_TOKEN         = os.getenv("BOT_TOKEN", "")
OWNER_CHAT_ID     = int(os.getenv("OWNER_CHAT_ID", "0"))
DATABASE_URL      = os.getenv("DATABASE_URL", "")
PORT              = int(os.getenv("PORT", "8080"))
SMTP_HOST         = os.getenv("SMTP_HOST", "")
SMTP_PORT_NUM     = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER         = os.getenv("SMTP_USER", "")
SMTP_PASS         = os.getenv("SMTP_PASS", "")
APP_URL           = os.getenv("APP_URL", "")
OPEN_REGISTRATION = os.getenv("OPEN_REGISTRATION", "true").lower() != "false"
TZ                = ZoneInfo("Asia/Jerusalem")
STATIC_DIR        = os.path.dirname(os.path.abspath(__file__))

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# Global Telegram Application (webhook mode — no polling, no Conflict on redeploy)
_tg_app: Optional[Application] = None

# ── Status / Source / Package constants ───────────────────────────────────────
ST_NEW          = "new"
ST_PRE_REMINDED = "pre_reminded"
ST_CALLED       = "called"
ST_WON          = "won"
ST_LOST         = "lost"
ST_FOLLOW_UP    = "follow_up"
ST_OLD          = "old"

STATUS_HEB = {
    ST_NEW: "חדש", ST_PRE_REMINDED: "תזכורת נשלחה", ST_CALLED: "עובד",
    ST_WON: "נסגר ✅", ST_LOST: "אבד ❌", ST_FOLLOW_UP: "המשך טיפול", ST_OLD: "ישן",
}
STATUS_EMOJI = {
    ST_NEW: "🆕", ST_PRE_REMINDED: "⏰", ST_CALLED: "📞",
    ST_WON: "✅", ST_LOST: "❌", ST_FOLLOW_UP: "🔄", ST_OLD: "📁",
}
SOURCE_HEB = {
    "instagram": "אינסטגרם",
    "facebook":  "פייסבוק",
    "tiktok":    "טיקטוק",
    "referral":  "המלצה",
    "website":   "אתר",
    "phone":     "טלפון נכנס",
    "other":     "אחר",
}
PACKAGE_HEB = {
    "basic":    "בסיסי",
    "standard": "סטנדרט",
    "premium":  "פרמיום",
    "vip":      "VIP",
    "custom":   "מותאם",
}

# ── DB pool ───────────────────────────────────────────────────────────────────
_pool: Optional[psycopg2.pool.ThreadedConnectionPool] = None

def get_pool() -> psycopg2.pool.ThreadedConnectionPool:
    global _pool
    if _pool is None:
        _pool = psycopg2.pool.ThreadedConnectionPool(1, 10, DATABASE_URL)
    return _pool

@contextmanager
def _db():
    pool = get_pool()
    conn = pool.getconn()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)

def _now() -> str:
    return datetime.now(TZ).strftime("%Y-%m-%dT%H:%M:%S")

def _parse_dt(s: str) -> Optional[datetime]:
    """Parse datetime string in either %Y-%m-%dT%H:%M:%S or %Y-%m-%dT%H:%M format."""
    if not s:
        return None
    try:
        return datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=TZ)
    except ValueError:
        pass
    try:
        return datetime.strptime(s[:16], "%Y-%m-%dT%H:%M").replace(tzinfo=TZ)
    except ValueError:
        pass
    return None

def ensure_db():
    with _db() as conn:
        cur = conn.cursor()
        # Leads table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS leads (
                id             TEXT PRIMARY KEY,
                name           TEXT NOT NULL,
                phone          TEXT NOT NULL,
                call_time      TEXT,
                status         TEXT DEFAULT 'new',
                sale_amount    REAL DEFAULT 0,
                notes          TEXT DEFAULT '',
                created_at     TEXT,
                updated_at     TEXT,
                follow_up_time TEXT,
                tags           TEXT DEFAULT '',
                created_by     TEXT DEFAULT '',
                source         TEXT DEFAULT '',
                package        TEXT DEFAULT '',
                user_id        TEXT DEFAULT ''
            )
        """)
        # Add new columns to existing databases
        for col, default in [
            ("tags", "''"), ("created_by", "''"), ("source", "''"),
            ("package", "''"), ("user_id", "''"),
        ]:
            cur.execute(f"ALTER TABLE leads ADD COLUMN IF NOT EXISTS {col} TEXT DEFAULT {default}")
        # Users table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id            TEXT PRIMARY KEY,
                username      TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                display_name  TEXT DEFAULT '',
                created_at    TEXT,
                is_admin      BOOLEAN DEFAULT FALSE
            )
        """)
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_admin BOOLEAN DEFAULT FALSE")
        # Sessions table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sessions (
                token      TEXT PRIMARY KEY,
                user_id    TEXT NOT NULL,
                created_at TEXT,
                expires_at TEXT
            )
        """)
        # Password reset tokens
        cur.execute("""
            CREATE TABLE IF NOT EXISTS reset_tokens (
                token      TEXT PRIMARY KEY,
                user_id    TEXT NOT NULL,
                code       TEXT NOT NULL,
                created_at TEXT,
                expires_at TEXT,
                used       BOOLEAN DEFAULT FALSE
            )
        """)
        # Push notification subscriptions
        cur.execute("""
            CREATE TABLE IF NOT EXISTS push_subscriptions (
                id           TEXT PRIMARY KEY,
                user_id      TEXT NOT NULL,
                subscription TEXT NOT NULL,
                created_at   TEXT
            )
        """)
        # App settings (VAPID keys, etc.)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)
    logger.info("DB ✓")

# ── Auth helpers ──────────────────────────────────────────────────────────────
def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    key  = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 100_000)
    return f"{salt}:{key.hex()}"

def verify_password(password: str, stored: str) -> bool:
    try:
        salt, key_hex = stored.split(":", 1)
        key = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 100_000)
        return _hmac.compare_digest(key.hex(), key_hex)
    except Exception:
        return False

# ── User / Session DB ─────────────────────────────────────────────────────────
def db_count_users() -> int:
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM users")
        return cur.fetchone()[0]

def db_get_user_by_username(username: str) -> Optional[dict]:
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM users WHERE username=%s", (username.lower().strip(),))
        row = cur.fetchone()
    return dict(row) if row else None

def db_create_user(username: str, password: str, display_name: str = "") -> str:
    uid      = str(uuid.uuid4())[:8].upper()
    is_admin = db_count_users() == 0  # First user is always admin
    with _db() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO users (id, username, password_hash, display_name, created_at, is_admin) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (uid, username.lower().strip(), hash_password(password),
             display_name or username, _now(), is_admin)
        )
    return uid

def db_create_session(user_id: str) -> str:
    token   = secrets.token_urlsafe(32)
    expires = (datetime.now(TZ) + timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%S")
    with _db() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO sessions (token, user_id, created_at, expires_at) VALUES (%s,%s,%s,%s)",
            (token, user_id, _now(), expires)
        )
    return token

def db_get_user_by_token(token: str) -> Optional[dict]:
    now = _now()
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT u.id, u.username, u.display_name, u.created_at, u.is_admin
            FROM users u
            JOIN sessions s ON s.user_id = u.id
            WHERE s.token = %s AND s.expires_at > %s
        """, (token, now))
        row = cur.fetchone()
    return dict(row) if row else None

def db_delete_session(token: str):
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM sessions WHERE token=%s", (token,))

def db_list_users() -> list:
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT u.id, u.username, u.display_name, u.created_at, u.is_admin,
                   COUNT(l.id) AS lead_count
            FROM users u
            LEFT JOIN leads l ON l.user_id = u.id
            GROUP BY u.id, u.username, u.display_name, u.created_at, u.is_admin
            ORDER BY u.created_at
        """)
        return [dict(r) for r in cur.fetchall()]

def db_get_first_user_id() -> Optional[str]:
    """Get the first created user's ID — used for Telegram-added leads."""
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id FROM users ORDER BY created_at LIMIT 1")
        row = cur.fetchone()
    return row[0] if row else None

# ── Password reset ────────────────────────────────────────────────────────────
def db_create_reset_token(user_id: str) -> str:
    """Generate a 6-digit reset code, store it, return the code."""
    code    = str(secrets.randbelow(900000) + 100000)  # always 6 digits
    token   = secrets.token_urlsafe(32)
    expires = (datetime.now(TZ) + timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M:%S")
    with _db() as conn:
        cur = conn.cursor()
        cur.execute(
            "UPDATE reset_tokens SET used=TRUE WHERE user_id=%s AND used=FALSE", (user_id,)
        )
        cur.execute(
            "INSERT INTO reset_tokens (token, user_id, code, created_at, expires_at) "
            "VALUES (%s,%s,%s,%s,%s)",
            (token, user_id, code, _now(), expires)
        )
    return code

def db_verify_reset_code(username: str, code: str) -> Optional[str]:
    """Verify a reset code; return user_id if valid, None otherwise."""
    now  = _now()
    user = db_get_user_by_username(username)
    if not user:
        return None
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT * FROM reset_tokens
            WHERE user_id=%s AND code=%s AND used=FALSE AND expires_at > %s
            ORDER BY created_at DESC LIMIT 1
        """, (user["id"], code, now))
        row = cur.fetchone()
        if not row:
            return None
        cur.execute("UPDATE reset_tokens SET used=TRUE WHERE token=%s", (row["token"],))
    return user["id"]

def db_update_password(user_id: str, new_password: str):
    with _db() as conn:
        cur = conn.cursor()
        cur.execute(
            "UPDATE users SET password_hash=%s WHERE id=%s",
            (hash_password(new_password), user_id)
        )

# ── Push subscriptions ────────────────────────────────────────────────────────
def db_get_setting(key: str) -> Optional[str]:
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT value FROM settings WHERE key=%s", (key,))
        row = cur.fetchone()
    return row[0] if row else None

def db_set_setting(key: str, value: str):
    with _db() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO settings(key,value) VALUES(%s,%s) "
            "ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value",
            (key, value)
        )

# ── VAPID / Web Push ──────────────────────────────────────────────────────────

_vapid_priv: Optional[str] = None
_vapid_pub:  Optional[str] = None

def _generate_vapid_keys():
    """Generate VAPID key pair. Private key stored as base64url raw bytes (most compatible)."""
    priv = ec.generate_private_key(ec.SECP256R1())
    # Store private key as base64url-encoded raw 32-byte integer
    priv_int = priv.private_numbers().private_value
    priv_b64 = base64.urlsafe_b64encode(priv_int.to_bytes(32, 'big')).rstrip(b'=').decode()
    pub_bytes = priv.public_key().public_bytes(
        serialization.Encoding.X962,
        serialization.PublicFormat.UncompressedPoint
    )
    pub_b64 = base64.urlsafe_b64encode(pub_bytes).rstrip(b'=').decode()
    return priv_b64, pub_b64

def _to_vapid_b64url(key: str) -> str:
    """Normalise any private key format → base64url raw bytes (pywebpush Vapid.from_string)."""
    if not key:
        return key
    key = key.strip()
    if key.startswith("-----"):
        # PEM (PKCS8 or EC) — extract raw private integer
        try:
            from cryptography.hazmat.primitives.serialization import load_pem_private_key
            priv_key = load_pem_private_key(key.encode(), password=None)
            priv_int = priv_key.private_numbers().private_value
            converted = base64.urlsafe_b64encode(priv_int.to_bytes(32, 'big')).rstrip(b'=').decode()
            logger.info("Converted PEM VAPID key → base64url format")
            return converted
        except Exception as e:
            logger.error(f"VAPID key conversion failed: {e}")
            return key
    return key  # already base64url

def get_vapid_keys() -> tuple:
    global _vapid_priv, _vapid_pub
    if _vapid_priv and _vapid_pub:
        return _vapid_priv, _vapid_pub
    # Env vars first
    p = os.getenv("VAPID_PRIVATE_KEY", "")
    q = os.getenv("VAPID_PUBLIC_KEY", "")
    if p and q:
        p = _to_vapid_b64url(p)
        _vapid_priv, _vapid_pub = p, q
        return p, q
    # DB
    p = db_get_setting("vapid_private_key") or ""
    q = db_get_setting("vapid_public_key") or ""
    if p and q:
        p = _to_vapid_b64url(p)
        if not p.startswith("-----"):  # converted successfully
            db_set_setting("vapid_private_key", p)  # save back in correct format
        _vapid_priv, _vapid_pub = p, q
        return p, q
    # Generate fresh keys
    if not WEBPUSH_AVAILABLE:
        return "", ""
    p, q = _generate_vapid_keys()
    db_set_setting("vapid_private_key", p)
    db_set_setting("vapid_public_key", q)
    _vapid_priv, _vapid_pub = p, q
    logger.info(f"VAPID keys generated. Public: {q[:20]}...")
    return p, q

def db_get_push_subscriptions_for_user(user_id: str) -> list:
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM push_subscriptions WHERE user_id=%s", (user_id,))
        return [dict(r) for r in cur.fetchall()]

def db_get_all_push_subscriptions() -> list:
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM push_subscriptions")
        return [dict(r) for r in cur.fetchall()]

def db_delete_push_subscription(sub_id: str):
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM push_subscriptions WHERE id=%s", (sub_id,))

def _send_push_one(sub_record: dict, title: str, body: str, url: str = "/", phone: str = "") -> bool:
    if not WEBPUSH_AVAILABLE:
        return False
    try:
        priv, pub = get_vapid_keys()
        if not priv:
            logger.warning("No VAPID private key — push skipped")
            return False
        sub_info = json.loads(sub_record["subscription"])
        payload = {"title": title, "body": body, "url": url}
        if phone:
            payload["phone"] = phone
        _webpush(
            subscription_info=sub_info,
            data=json.dumps(payload),
            vapid_private_key=priv,
            vapid_claims={"sub": "mailto:admin@bar-lead-bot.onrender.com"},
        )
        logger.info(f"Push sent OK: {title}")
        return True
    except WebPushException as ex:
        resp = getattr(ex, "response", None)
        status = resp.status_code if resp else "N/A"
        body_text = ""
        if resp:
            try: body_text = resp.text[:200]
            except: pass
        logger.error(f"WebPushException [{status}] body={body_text}: {ex}")
        if resp and resp.status_code in (404, 410):
            db_delete_push_subscription(sub_record["id"])
            logger.info("Deleted stale push subscription")
        raise  # re-raise so test endpoint can show the real error
    except Exception as e:
        logger.error(f"push error: {e}", exc_info=True)
        raise

def push_to_user(user_id: str, title: str, body: str, url: str = "/", phone: str = ""):
    if not WEBPUSH_AVAILABLE:
        return
    subs = db_get_push_subscriptions_for_user(user_id)
    logger.info(f"push_to_user: user={user_id} subs={len(subs)} title={title}")
    for sub in subs:
        try:
            _send_push_one(sub, title, body, url, phone)
        except Exception as e:
            logger.error(f"push_to_user send failed: {e}")

def push_to_all(title: str, body: str, url: str = "/", phone: str = ""):
    if not WEBPUSH_AVAILABLE:
        return
    subs = db_get_all_push_subscriptions()
    logger.info(f"push_to_all: subs={len(subs)} title={title}")
    for sub in subs:
        _send_push_one(sub, title, body, url, phone)

def _send_tg(html_msg: str) -> None:
    """Synchronous fire-and-forget Telegram notification (never raises)."""
    if not BOT_TOKEN or not OWNER_CHAT_ID:
        return
    try:
        payload = json.dumps({
            "chat_id": OWNER_CHAT_ID,
            "text":    html_msg,
            "parse_mode": "HTML",
        }).encode()
        req = _ureq.Request(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        _ureq.urlopen(req, timeout=5)
    except Exception as e:
        logger.warning(f"_send_tg failed (non-critical): {e}")

# ─────────────────────────────────────────────────────────────────────────────

def db_save_push_subscription(user_id: str, subscription_json: str):
    sub_id = str(uuid.uuid4())[:8]
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM push_subscriptions WHERE user_id=%s", (user_id,))
        cur.execute(
            "INSERT INTO push_subscriptions (id, user_id, subscription, created_at) "
            "VALUES (%s,%s,%s,%s)",
            (sub_id, user_id, subscription_json, _now())
        )

# ── Email helper ──────────────────────────────────────────────────────────────
def send_email(to_email: str, subject: str, body_html: str) -> bool:
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASS:
        logger.warning("SMTP not configured — email not sent")
        return False
    try:
        msg            = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = SMTP_USER
        msg["To"]      = to_email
        msg.attach(MIMEText(body_html, "html", "utf-8"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT_NUM, timeout=10) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, [to_email], msg.as_string())
        return True
    except Exception as e:
        logger.error(f"send_email: {e}")
        return False

# ── Lead DB operations ────────────────────────────────────────────────────────
def db_get_lead(lid: str) -> Optional[dict]:
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM leads WHERE id=%s", (lid,))
        row = cur.fetchone()
    return dict(row) if row else None

def db_get_all(user_id: Optional[str] = None) -> list:
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if user_id:
            cur.execute("SELECT * FROM leads WHERE user_id=%s ORDER BY call_time", (user_id,))
        else:
            cur.execute("SELECT * FROM leads ORDER BY call_time")
        return [dict(r) for r in cur.fetchall()]

def db_get_active(user_id: Optional[str] = None) -> list:
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if user_id:
            cur.execute(
                "SELECT * FROM leads WHERE status IN (%s,%s,%s,%s) AND user_id=%s ORDER BY call_time",
                (ST_NEW, ST_PRE_REMINDED, ST_CALLED, ST_FOLLOW_UP, user_id)
            )
        else:
            cur.execute(
                "SELECT * FROM leads WHERE status IN (%s,%s,%s,%s) ORDER BY call_time",
                (ST_NEW, ST_PRE_REMINDED, ST_CALLED, ST_FOLLOW_UP)
            )
        return [dict(r) for r in cur.fetchall()]

def db_get_today(user_id: Optional[str] = None) -> list:
    today = datetime.now(TZ).strftime("%Y-%m-%d")
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if user_id:
            cur.execute(
                "SELECT * FROM leads WHERE call_time LIKE %s AND user_id=%s ORDER BY call_time",
                (f"{today}%", user_id)
            )
        else:
            cur.execute(
                "SELECT * FROM leads WHERE call_time LIKE %s ORDER BY call_time", (f"{today}%",)
            )
        return [dict(r) for r in cur.fetchall()]

def db_get_old(user_id: Optional[str] = None) -> list:
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if user_id:
            cur.execute(
                "SELECT * FROM leads WHERE status=%s AND user_id=%s ORDER BY updated_at DESC",
                (ST_OLD, user_id)
            )
        else:
            cur.execute("SELECT * FROM leads WHERE status=%s ORDER BY updated_at DESC", (ST_OLD,))
        return [dict(r) for r in cur.fetchall()]

def db_search(query: str, user_id: Optional[str] = None) -> list:
    q = f"%{query.lower()}%"
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if user_id:
            cur.execute(
                "SELECT * FROM leads WHERE (LOWER(name) LIKE %s OR phone LIKE %s) AND user_id=%s",
                (q, q, user_id)
            )
        else:
            cur.execute(
                "SELECT * FROM leads WHERE LOWER(name) LIKE %s OR phone LIKE %s", (q, q)
            )
        return [dict(r) for r in cur.fetchall()]

def db_add_lead(name: str, phone: str, call_dt, user_id: str = "") -> str:
    lid       = str(uuid.uuid4())[:8].upper()
    now       = _now()
    call_time = call_dt.strftime("%Y-%m-%dT%H:%M:%S") if call_dt else None
    if not user_id:
        user_id = db_get_first_user_id() or ""
    with _db() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO leads (id, name, phone, call_time, status, created_at, updated_at, user_id) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (lid, name, phone, call_time, ST_NEW, now, now, user_id)
        )
    return lid

def db_add_lead_full(name: str, phone: str, call_time: Optional[str],
                     notes: str, tags: str, created_by: str = "",
                     source: str = "", package: str = "", user_id: str = "") -> str:
    lid = str(uuid.uuid4())[:8].upper()
    now = _now()
    if not user_id:
        user_id = db_get_first_user_id() or ""
    with _db() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO leads "
            "(id, name, phone, call_time, status, notes, tags, created_by, "
            "source, package, user_id, created_at, updated_at) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (lid, name, phone, call_time, ST_NEW, notes or "", tags or "", created_by,
             source or "", package or "", user_id, now, now)
        )
    return lid

def db_update_lead(lid: str, **kw):
    kw["updated_at"] = _now()
    if kw.get("status") in (ST_WON, ST_LOST, ST_OLD) and "follow_up_time" not in kw:
        kw["follow_up_time"] = None
    cols = ", ".join(f"{k}=%s" for k in kw)
    vals = list(kw.values()) + [lid]
    with _db() as conn:
        cur = conn.cursor()
        cur.execute(f"UPDATE leads SET {cols} WHERE id=%s", vals)

def db_delete_lead(lid: str):
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM leads WHERE id=%s", (lid,))

def db_get_leads_pre_reminder() -> list:
    """Return leads that need a reminder now.
    Window: call_time is between 30 min PAST and 20 min FUTURE (catches missed reminders if server was sleeping)."""
    now = datetime.now(TZ)
    out = []
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM leads WHERE status=%s AND call_time IS NOT NULL", (ST_NEW,))
        rows = cur.fetchall()
    for r in rows:
        dt   = _parse_dt(str(r.get("call_time", "")))
        if dt is None:
            continue
        diff = (dt - now).total_seconds() / 60
        # -30 catches cases where the server was asleep and woke up after the call time
        if -30 <= diff <= 20:
            out.append(dict(r))
    return out

def db_get_leads_post_call() -> list:
    now = datetime.now(TZ)
    out = []
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT * FROM leads WHERE status IN (%s,%s) AND call_time IS NOT NULL",
            (ST_NEW, ST_PRE_REMINDED)
        )
        rows = cur.fetchall()
    for r in rows:
        dt   = _parse_dt(str(r.get("call_time", "")))
        if dt is None:
            continue
        diff = (now - dt).total_seconds() / 60
        if diff >= 30:
            out.append(dict(r))
    return out

def db_get_leads_followup_due() -> list:
    now_str = _now()
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT * FROM leads WHERE status=%s AND follow_up_time IS NOT NULL AND follow_up_time <= %s",
            (ST_FOLLOW_UP, now_str)
        )
        return [dict(r) for r in cur.fetchall()]

def db_get_stale_leads() -> list:
    threshold = (datetime.now(TZ) - timedelta(days=3)).strftime("%Y-%m-%dT%H:%M:%S")
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT * FROM leads WHERE status=%s AND updated_at <= %s",
            (ST_CALLED, threshold)
        )
        return [dict(r) for r in cur.fetchall()]

def db_stats_since(since: str, user_id: Optional[str] = None) -> dict:
    uid_filter = "AND user_id=%s" if user_id else ""
    uid_params = [user_id] if user_id else []
    with _db() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT COUNT(*) FROM leads WHERE created_at >= %s {uid_filter}",
            [since] + uid_params
        )
        total = cur.fetchone()[0]
        cur.execute(
            f"SELECT COUNT(*), COALESCE(SUM(sale_amount),0) FROM leads "
            f"WHERE status=%s AND created_at >= %s {uid_filter}",
            [ST_WON, since] + uid_params
        )
        won = cur.fetchone()
        cur.execute(
            f"SELECT COUNT(*) FROM leads WHERE status=%s AND created_at >= %s {uid_filter}",
            [ST_LOST, since] + uid_params
        )
        lost = cur.fetchone()[0]
        cur.execute(
            f"SELECT COUNT(*) FROM leads WHERE status IN (%s,%s,%s,%s) AND created_at >= %s {uid_filter}",
            [ST_NEW, ST_PRE_REMINDED, ST_CALLED, ST_FOLLOW_UP, since] + uid_params
        )
        active = cur.fetchone()[0]
    return {
        "total":      total,
        "won_count":  won[0],
        "won_amount": float(won[1]),
        "lost":       lost,
        "active":     active,
    }

def db_source_counts(user_id: Optional[str] = None) -> dict:
    """Return leads count grouped by source for a user."""
    with _db() as conn:
        cur = conn.cursor()
        if user_id:
            cur.execute(
                "SELECT COALESCE(NULLIF(source,''),'other'), COUNT(*) "
                "FROM leads WHERE user_id=%s GROUP BY source",
                (user_id,)
            )
        else:
            cur.execute(
                "SELECT COALESCE(NULLIF(source,''),'other'), COUNT(*) FROM leads GROUP BY source"
            )
        return {row[0]: row[1] for row in cur.fetchall()}

def db_source_conversion(user_id: Optional[str] = None) -> dict:
    """Per-source totals and won counts, for conversion-rate analysis."""
    with _db() as conn:
        cur = conn.cursor()
        if user_id:
            cur.execute(
                "SELECT COALESCE(NULLIF(source,''),'other') AS src, "
                "COUNT(*) AS total, "
                "COUNT(*) FILTER (WHERE status=%s) AS won "
                "FROM leads WHERE user_id=%s GROUP BY src",
                (ST_WON, user_id)
            )
        else:
            cur.execute(
                "SELECT COALESCE(NULLIF(source,''),'other') AS src, "
                "COUNT(*) AS total, "
                "COUNT(*) FILTER (WHERE status=%s) AS won "
                "FROM leads GROUP BY src",
                (ST_WON,)
            )
        return {row[0]: {"total": row[1], "won": row[2]} for row in cur.fetchall()}

def db_calendar_leads(user_id: str, from_dt: str, to_dt: str) -> list:
    """Get leads with call_time in [from_dt, to_dt] for calendar view."""
    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT * FROM leads WHERE user_id=%s "
            "AND call_time IS NOT NULL AND call_time >= %s AND call_time <= %s "
            "ORDER BY call_time",
            (user_id, from_dt, to_dt)
        )
        return [dict(r) for r in cur.fetchall()]

# ─────────────────────────────────────────────────────────────────────────────
#  TELEGRAM BOT
# ─────────────────────────────────────────────────────────────────────────────

ASK_NAME, ASK_PHONE, ASK_TIME = range(3)

MENU_BUTTONS = {
    "➕ הוספת ליד", "📋 לידים ישנים", "📊 סיכום", "⏰ תזכורות פעילות",
    "📅 היום", "🔍 חיפוש ליד", "📤 יצוא Excel", "✏️ עריכת ליד",
}

MAIN_MENU = ReplyKeyboardMarkup(
    [
        [KeyboardButton("➕ הוספת ליד"),      KeyboardButton("📋 לידים ישנים")],
        [KeyboardButton("📊 סיכום"),           KeyboardButton("⏰ תזכורות פעילות")],
        [KeyboardButton("📅 היום"),            KeyboardButton("🔍 חיפוש ליד")],
        [KeyboardButton("📤 יצוא Excel"),      KeyboardButton("✏️ עריכת ליד")],
    ],
    resize_keyboard=True,
    is_persistent=True,
)

def parse_time(text: str):
    now = datetime.now(TZ)
    t   = text.strip()
    if "מחר" in t:
        base = (now + timedelta(days=1)).replace(second=0, microsecond=0)
        for w in t.split():
            if ":" in w:
                try:
                    h, m = map(int, w.split(":"))
                    return base.replace(hour=h, minute=m)
                except Exception:
                    pass
        return base.replace(hour=9, minute=0)
    if "שעות" in t or "שעה" in t:
        for w in t.split():
            if w.isdigit():
                return now + timedelta(hours=int(w))
        return now + timedelta(hours=1)
    if "דקות" in t or "דקה" in t:
        for w in t.split():
            if w.isdigit():
                return now + timedelta(minutes=int(w))
        return now + timedelta(minutes=30)
    parts = t.split()
    if len(parts) == 2 and "/" in parts[0] and ":" in parts[1]:
        try:
            day, month = map(int, parts[0].split("/"))
            h, m       = map(int, parts[1].split(":"))
            dt = now.replace(month=month, day=day, hour=h, minute=m, second=0, microsecond=0)
            if dt < now:
                dt = dt.replace(year=now.year + 1)
            return dt
        except Exception:
            pass
    if ":" in t:
        try:
            h, m = map(int, t.split(":"))
            dt   = now.replace(hour=h, minute=m, second=0, microsecond=0)
            return dt if dt > now else dt + timedelta(days=1)
        except Exception:
            pass
    return None

def validate_phone(phone: str) -> bool:
    digits = "".join(c for c in phone if c.isdigit())
    if digits.startswith("972"):
        digits = "0" + digits[3:]
    return len(digits) == 10 and digits.startswith("05")

def outcome_kb(lid: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ נסגר! 🎉",        callback_data=f"out_{lid}_won")],
        [InlineKeyboardButton("⏰ נדחה — קבע זמן",  callback_data=f"out_{lid}_snooze"),
         InlineKeyboardButton("🔄 ליד ישן",          callback_data=f"out_{lid}_old")],
        [InlineKeyboardButton("❌ לא רלוונטי",       callback_data=f"out_{lid}_lost")],
    ])

def snooze_kb(lid: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("בעוד שעה ⏰",  callback_data=f"snz_{lid}_1h"),
         InlineKeyboardButton("מחר 9:00 🌅",  callback_data=f"snz_{lid}_tomorrow")],
        [InlineKeyboardButton("קבע ידנית ✏️", callback_data=f"snz_{lid}_manual")],
    ])

def edit_field_kb(lid: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👤 שינוי שם",        callback_data=f"edf_{lid}_name")],
        [InlineKeyboardButton("📱 שינוי טלפון",      callback_data=f"edf_{lid}_phone")],
        [InlineKeyboardButton("📅 שינוי שעת שיחה",  callback_data=f"edf_{lid}_time")],
        [InlineKeyboardButton("🗑 מחיקת ליד",        callback_data=f"edf_{lid}_delete")],
    ])

def build_excel(user_id: Optional[str] = None) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "לידים"
    hf = Font(bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor="2B579A")
    ca = Alignment(horizontal="center")
    col_labels = ["מזהה", "שם", "טלפון", "שעת שיחה", "סטטוס", "מקור",
                  "חבילה", "סכום מכירה", "הערות", "תגיות", "נוצר ע״י", "נוצר"]
    col_keys   = ["id", "name", "phone", "call_time", "status", "source",
                  "package", "sale_amount", "notes", "tags", "created_by", "created_at"]
    for ci, label in enumerate(col_labels, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = hf; cell.fill = hb; cell.alignment = ca
    for ri, row in enumerate(db_get_all(user_id), 2):
        for ci, key in enumerate(col_keys, 1):
            val = row.get(key, "")
            if key == "status":
                val = STATUS_HEB.get(str(val), str(val))
            elif key == "source":
                val = SOURCE_HEB.get(str(val), str(val) or "")
            elif key == "package":
                val = PACKAGE_HEB.get(str(val), str(val) or "")
            elif key in ("call_time", "created_at") and val:
                dt = _parse_dt(str(val))
                if dt:
                    val = dt.strftime("%d/%m/%Y %H:%M")
            elif key == "sale_amount" and val:
                try:
                    val = f"₪{float(val):,.0f}"
                except Exception:
                    pass
            ws.cell(row=ri, column=ci, value=str(val) if val else "")
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

async def add_start(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await u.message.reply_text(
        "👤 *מה שם הלקוח?*\n\n_שלח /cancel לביטול_", parse_mode="Markdown")
    return ASK_NAME

async def got_name(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = u.message.text.strip()
    if name in MENU_BUTTONS:
        return await cancel_add(u, ctx)
    ctx.user_data["name"] = name
    await u.message.reply_text(
        f"✅ שם: *{name}*\n\n📱 *מה מספר הטלפון?*\n_לדוגמא: 0501234567_",
        parse_mode="Markdown")
    return ASK_PHONE

async def got_phone(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    phone = u.message.text.strip()
    if phone in MENU_BUTTONS:
        return await cancel_add(u, ctx)
    if not validate_phone(phone):
        await u.message.reply_text(
            "❌ מספר טלפון לא תקין!\nנא להכניס מספר ישראלי: *0501234567*",
            parse_mode="Markdown")
        return ASK_PHONE
    ctx.user_data["phone"] = phone
    await u.message.reply_text(
        f"✅ טלפון: *{phone}*\n\n"
        "📅 *מתי השיחה?*\n"
        "_לדוגמא:_\n"
        "`16:30` — היום בשעה 16:30\n"
        "`מחר 10:00` — מחר בעשר\n"
        "`25/06 15:30` — תאריך ספציפי\n"
        "`בעוד שעה` — עוד שעה מעכשיו",
        parse_mode="Markdown")
    return ASK_TIME

async def got_time(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = u.message.text.strip()
    if text in MENU_BUTTONS:
        return await cancel_add(u, ctx)
    call_dt = parse_time(text)
    if not call_dt:
        await u.message.reply_text(
            "❓ לא הבנתי. נסה:\n`16:30` | `מחר 10:00` | `25/06 15:30`",
            parse_mode="Markdown")
        return ASK_TIME
    name  = ctx.user_data["name"]
    phone = ctx.user_data["phone"]
    lid   = db_add_lead(name, phone, call_dt)
    ctx.user_data.clear()
    await u.message.reply_text(
        f"🎯 *ליד נוסף בהצלחה!*\n\n"
        f"👤 *{name}*\n"
        f"📱 {phone}\n"
        f"📅 שיחה: *{call_dt.strftime('%d/%m/%Y בשעה %H:%M')}*\n\n"
        f"⏰ תקבל תזכורת 20 דקות לפני",
        parse_mode="Markdown", reply_markup=MAIN_MENU)
    return ConversationHandler.END

async def cancel_add(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    if u.message:
        await u.message.reply_text("❌ הוספת ליד בוטלה.", reply_markup=MAIN_MENU)
    return ConversationHandler.END

async def cmd_start(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await u.message.reply_text(
        "👋 *שלום! בוט ניהול הלידים שלך* 🍸\n\nבחר פעולה:",
        parse_mode="Markdown", reply_markup=MAIN_MENU)

async def show_today(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    leads  = db_get_today()
    active = [l for l in leads if l.get("status") in (ST_NEW, ST_PRE_REMINDED, ST_CALLED, ST_FOLLOW_UP)]
    if not active:
        await u.message.reply_text("📅 אין שיחות מתוכננות להיום! 🎉", reply_markup=MAIN_MENU)
        return
    lines = [f"📅 *שיחות להיום — {datetime.now(TZ).strftime('%d/%m/%Y')} ({len(active)}):*\n"]
    for ld in active:
        dt = _parse_dt(str(ld.get("call_time", "")))
        t  = dt.strftime("%H:%M") if dt else "—"
        lines.append(
            f"🕐 *{t}* — {STATUS_EMOJI.get(ld.get('status'), '❓')} "
            f"*{ld['name']}* | 📱{ld.get('phone', '—')}"
        )
    await u.message.reply_text("\n".join(lines), parse_mode="Markdown", reply_markup=MAIN_MENU)

async def show_old_leads(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    old = db_get_old()
    if not old:
        await u.message.reply_text("📁 אין לידים ישנים.", reply_markup=MAIN_MENU)
        return
    await u.message.reply_text(f"📁 *לידים ישנים ({len(old)}):*", parse_mode="Markdown")
    for ld in old:
        ct       = str(ld.get("call_time", "") or "")
        date_str = ct[:10].replace("-", "/") if ct else "—"
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📞 לחזור אליו", callback_data=f"recall_{ld['id']}"),
             InlineKeyboardButton("✅ נסגר",        callback_data=f"out_{ld['id']}_won")],
            [InlineKeyboardButton("❌ לא רלוונטי", callback_data=f"out_{ld['id']}_lost")],
        ])
        await u.message.reply_text(
            f"📁 *{ld['name']}*\n📱 {ld.get('phone', '—')}\n📅 שיחה אחרונה: {date_str}",
            reply_markup=kb, parse_mode="Markdown")

async def show_reminders(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    active = db_get_active()
    if not active:
        await u.message.reply_text("✅ אין תזכורות פעילות!", reply_markup=MAIN_MENU)
        return
    lines = [f"⏰ *תזכורות פעילות ({len(active)}):*\n"]
    for ld in active:
        ct = str(ld.get("call_time") or ld.get("follow_up_time") or "")
        t  = ""
        if ct:
            dt = _parse_dt(ct)
            if dt:
                t = f" | ⏰ {dt.strftime('%d/%m %H:%M')}"
        lines.append(
            f"{STATUS_EMOJI.get(ld.get('status'), '❓')} *{ld['name']}* "
            f"— 📱{ld.get('phone', '—')}{t}"
        )
    await u.message.reply_text("\n".join(lines), parse_mode="Markdown", reply_markup=MAIN_MENU)

async def show_summary(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    now     = datetime.now(TZ)
    weekly  = db_stats_since((now - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S"))
    monthly = db_stats_since(now.replace(day=1, hour=0, minute=0, second=0).strftime("%Y-%m-%dT%H:%M:%S"))
    rate    = lambda s: f"{s['won_count']/s['total']*100:.0f}%" if s["total"] else "—"
    await u.message.reply_text(
        "📊 *סיכום ביצועים*\n\n"
        f"📅 *שבוע אחרון:*\n"
        f"  לידים: {weekly['total']} | ✅ נסגרו: {weekly['won_count']} | ❌ אבדו: {weekly['lost']}\n"
        f"  💰 הכנסות: ₪{weekly['won_amount']:,.0f} | המרה: {rate(weekly)}\n\n"
        f"📆 *החודש הנוכחי:*\n"
        f"  לידים: {monthly['total']} | ✅ נסגרו: {monthly['won_count']} | ❌ אבדו: {monthly['lost']}\n"
        f"  💰 הכנסות: ₪{monthly['won_amount']:,.0f} | המרה: {rate(monthly)}\n\n"
        f"🟢 פעילים: {len(db_get_active())} | 📁 ישנים: {len(db_get_old())}",
        parse_mode="Markdown", reply_markup=MAIN_MENU)

async def start_search(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["state"] = "searching"
    await u.message.reply_text("🔍 *חיפוש ליד*\n\nהכנס שם או מספר טלפון:", parse_mode="Markdown")

async def export_excel_bot(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await u.message.reply_text("⏳ מכין קובץ Excel...", reply_markup=MAIN_MENU)
    try:
        buf   = build_excel()
        fname = f"leads_{datetime.now(TZ).strftime('%d%m%Y_%H%M')}.xlsx"
        await u.message.reply_document(
            document=buf, filename=fname,
            caption=f"📊 כל הלידים — {datetime.now(TZ).strftime('%d/%m/%Y %H:%M')}")
    except Exception as e:
        logger.error(f"export_excel_bot: {e}")
        await u.message.reply_text(f"❌ שגיאה ביצירת קובץ: {e}")

async def start_edit(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    active = db_get_active()
    if not active:
        await u.message.reply_text("❌ אין לידים פעילים לעריכה.", reply_markup=MAIN_MENU)
        return
    buttons = []
    for ld in active[:10]:
        dt = _parse_dt(str(ld.get("call_time", "")))
        t  = f" {dt.strftime('%d/%m %H:%M')}" if dt else ""
        buttons.append([InlineKeyboardButton(
            f"{STATUS_EMOJI.get(ld.get('status'), '❓')} {ld['name']}{t}",
            callback_data=f"edit_{ld['id']}"
        )])
    await u.message.reply_text(
        "✏️ *בחר ליד לעריכה:*",
        reply_markup=InlineKeyboardMarkup(buttons), parse_mode="Markdown")

async def handle_text(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text  = u.message.text.strip()
    state = ctx.user_data.get("state")

    if state == "ask_amount":
        lid = ctx.user_data.get("lid")
        try:
            amount = float(text.replace("₪", "").replace(",", ""))
            ld     = db_get_lead(lid)
            db_update_lead(lid, status=ST_WON, sale_amount=amount)
            ctx.user_data.clear()
            await u.message.reply_text(
                f"🎉 *כל הכבוד!*\n*{ld['name']}* — עסקה נסגרה\n💰 *₪{amount:,.0f}*",
                parse_mode="Markdown", reply_markup=MAIN_MENU)
        except ValueError:
            await u.message.reply_text(
                "❓ הכנס סכום בשקלים (לדוגמא: `3500`)", parse_mode="Markdown")
        return

    if state in ("ask_snooze", "ask_recall"):
        lid = ctx.user_data.get("lid")
        dt  = parse_time(text)
        if dt:
            if state == "ask_recall":
                db_update_lead(lid, status=ST_NEW,
                               call_time=dt.strftime("%Y-%m-%dT%H:%M:%S"),
                               follow_up_time=None)
            else:
                db_update_lead(lid, status=ST_FOLLOW_UP,
                               follow_up_time=dt.strftime("%Y-%m-%dT%H:%M:%S"))
            ctx.user_data.clear()
            await u.message.reply_text(
                f"⏰ תזכורת נקבעה ל-*{dt.strftime('%d/%m/%Y %H:%M')}* ✓",
                parse_mode="Markdown", reply_markup=MAIN_MENU)
        else:
            await u.message.reply_text(
                "❓ נסה: `16:30` | `מחר 10:00` | `25/06 15:00`", parse_mode="Markdown")
        return

    if state == "searching":
        results = db_search(text)
        ctx.user_data.clear()
        if not results:
            await u.message.reply_text(
                f"🔍 לא נמצאו לידים עבור: *{text}*",
                parse_mode="Markdown", reply_markup=MAIN_MENU)
            return
        await u.message.reply_text(
            f"🔍 *{len(results)} תוצאות עבור \"{text}\":*", parse_mode="Markdown")
        for ld in results[:8]:
            dt = _parse_dt(str(ld.get("call_time", "")))
            t  = f"\n📅 {dt.strftime('%d/%m/%Y %H:%M')}" if dt else ""
            kb = InlineKeyboardMarkup([[
                InlineKeyboardButton("✏️ עריכה",       callback_data=f"edit_{ld['id']}"),
                InlineKeyboardButton("📋 עדכן סטטוס",  callback_data=f"out_{ld['id']}_snooze"),
            ]])
            await u.message.reply_text(
                f"{STATUS_EMOJI.get(ld.get('status'), '❓')} *{ld['name']}*\n"
                f"📱 {ld.get('phone', '—')}{t}",
                parse_mode="Markdown", reply_markup=kb)
        return

    if state == "edit_value":
        lid   = ctx.user_data.get("lid")
        field = ctx.user_data.get("field")
        if field == "name":
            db_update_lead(lid, name=text)
            ctx.user_data.clear()
            await u.message.reply_text(
                f"✅ שם עודכן ל-*{text}*", parse_mode="Markdown", reply_markup=MAIN_MENU)
        elif field == "phone":
            if not validate_phone(text):
                await u.message.reply_text(
                    "❌ מספר לא תקין! נסה שוב (לדוגמא: 0501234567)", parse_mode="Markdown")
                return
            db_update_lead(lid, phone=text)
            ctx.user_data.clear()
            await u.message.reply_text(
                f"✅ טלפון עודכן ל-*{text}*", parse_mode="Markdown", reply_markup=MAIN_MENU)
        elif field == "time":
            dt = parse_time(text)
            if not dt:
                await u.message.reply_text(
                    "❓ נסה: `16:30` | `מחר 10:00` | `25/06 15:00`", parse_mode="Markdown")
                return
            db_update_lead(lid, call_time=dt.strftime("%Y-%m-%dT%H:%M:%S"))
            ctx.user_data.clear()
            await u.message.reply_text(
                f"✅ שעת שיחה עודכנה ל-*{dt.strftime('%d/%m/%Y %H:%M')}*",
                parse_mode="Markdown", reply_markup=MAIN_MENU)
        return

    ctx.user_data.clear()
    if "לידים ישנים" in text:
        await show_old_leads(u, ctx)
    elif "סיכום" in text:
        await show_summary(u, ctx)
    elif "תזכורות" in text:
        await show_reminders(u, ctx)
    elif "היום" in text:
        await show_today(u, ctx)
    elif "חיפוש" in text:
        await start_search(u, ctx)
    elif "Excel" in text or "יצוא" in text:
        await export_excel_bot(u, ctx)
    elif "עריכת ליד" in text:
        await start_edit(u, ctx)
    else:
        await u.message.reply_text("💬 לחץ על אחת האפשרויות בתפריט 👇", reply_markup=MAIN_MENU)

async def callback_handler(u: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = u.callback_query
    await q.answer()
    data = q.data

    if data.startswith("edit_"):
        lid = data.replace("edit_", "")
        ld  = db_get_lead(lid)
        if not ld:
            await q.edit_message_text("❌ ליד לא נמצא"); return
        dt = _parse_dt(str(ld.get("call_time", "")))
        t  = dt.strftime("%d/%m/%Y %H:%M") if dt else "—"
        await q.edit_message_text(
            f"✏️ *עריכת ליד*\n\n👤 {ld['name']}\n📱 {ld.get('phone', '—')}\n📅 {t}\n\n*מה לשנות?*",
            reply_markup=edit_field_kb(lid), parse_mode="Markdown")
        return

    if data.startswith("edf_"):
        _, lid, field = data.split("_", 2)
        ld = db_get_lead(lid)
        if not ld:
            await q.edit_message_text("❌ ליד לא נמצא"); return
        if field == "delete":
            db_delete_lead(lid)
            await q.edit_message_text(f"🗑 *{ld['name']}* נמחק.", parse_mode="Markdown"); return
        prompts = {
            "name":  f"👤 הכנס שם חדש עבור *{ld['name']}*:",
            "phone": f"📱 הכנס טלפון חדש עבור *{ld['name']}*:",
            "time":  f"📅 הכנס שעת שיחה חדשה עבור *{ld['name']}*:\n`16:30` | `מחר 10:00` | `25/06 15:00`",
        }
        ctx.user_data.update({"state": "edit_value", "lid": lid, "field": field})
        await q.edit_message_text(prompts.get(field, "?"), parse_mode="Markdown"); return

    if data.startswith("out_"):
        _, lid, action = data.split("_", 2)
        ld = db_get_lead(lid)
        if not ld:
            await q.edit_message_text("❌ ליד לא נמצא"); return
        if action == "won":
            ctx.user_data.update({"state": "ask_amount", "lid": lid})
            await q.edit_message_text(
                f"🎉 *כמה שילם {ld['name']}?*\n_(הכנס סכום בשקלים)_", parse_mode="Markdown")
        elif action == "lost":
            db_update_lead(lid, status=ST_LOST)
            await q.edit_message_text(f"❌ *{ld['name']}* סומן כלא רלוונטי.", parse_mode="Markdown")
        elif action == "snooze":
            await q.edit_message_text(
                f"⏰ *{ld['name']}* — מתי לחזור אליו?",
                reply_markup=snooze_kb(lid), parse_mode="Markdown")
        elif action == "old":
            db_update_lead(lid, status=ST_OLD)
            await q.edit_message_text(f"📁 *{ld['name']}* הועבר ללידים ישנים.", parse_mode="Markdown")
        return

    if data.startswith("snz_"):
        parts  = data.split("_")
        lid    = parts[1]
        option = "_".join(parts[2:])
        now    = datetime.now(TZ)
        if option == "1h":
            new_dt = now + timedelta(hours=1)
        elif option == "tomorrow":
            new_dt = (now + timedelta(days=1)).replace(hour=9, minute=0, second=0, microsecond=0)
        elif option == "manual":
            ctx.user_data.update({"state": "ask_snooze", "lid": lid})
            ld = db_get_lead(lid)
            await q.edit_message_text(
                f"⏰ מתי לחזור אל *{ld['name']}*?\n`16:30` | `מחר 10:00` | `בעוד שעה`",
                parse_mode="Markdown"); return
        else:
            return
        db_update_lead(lid, status=ST_FOLLOW_UP,
                       follow_up_time=new_dt.strftime("%Y-%m-%dT%H:%M:%S"))
        ld = db_get_lead(lid)
        await q.edit_message_text(
            f"⏰ *{ld['name']}* — אחזור ב-*{new_dt.strftime('%d/%m %H:%M')}* 👍",
            parse_mode="Markdown"); return

    if data.startswith("recall_"):
        lid = data.replace("recall_", "")
        ld  = db_get_lead(lid)
        ctx.user_data.update({"state": "ask_recall", "lid": lid})
        await q.edit_message_text(
            f"📞 מתי השיחה עם *{ld['name']}*?\n`16:30` | `מחר 10:00` | `25/06 15:00`",
            parse_mode="Markdown")

# ── Scheduled jobs ────────────────────────────────────────────────────────────

async def job_check_reminders(app):
    for ld in db_get_leads_pre_reminder():
        try:
            dt        = _parse_dt(str(ld.get("call_time", "")))
            time_str  = dt.strftime('%H:%M') if dt else "—"
            now       = datetime.now(TZ)
            mins_diff = int((dt - now).total_seconds() / 60) if dt else 20
            missed    = mins_diff < 0  # server was sleeping and we woke up late

            if missed:
                mins_ago = abs(mins_diff)
                tg_text  = (f"🚨 *תזכורת שפוספסה!*\n\n"
                            f"השרת היה כבוי — שיחה עם *{ld['name']}* הייתה לפני {mins_ago} דקות!\n"
                            f"📱 {ld.get('phone', '—')}\n🕐 שעה {time_str}")
                push_title = "🚨 תזכורת שפוספסה!"
                push_body  = f"{ld['name']} · שיחה הייתה בשעה {time_str}"
            else:
                tg_text  = (f"⏰ *שיחה בעוד {mins_diff} דקות!*\n\n"
                            f"👤 *{ld['name']}*\n"
                            f"📱 {ld.get('phone', '—')}\n"
                            f"🕐 שיחה בשעה *{time_str}*")
                push_title = f"⏰ שיחה בעוד {mins_diff} דקות"
                push_body  = f"{ld['name']} · {ld.get('phone','—')} בשעה {time_str}"

            # Web Push — primary notification channel
            uid = ld.get("user_id") or db_get_first_user_id()
            push_sent = False
            if uid:
                push_subs = db_get_push_subscriptions_for_user(uid)
                if push_subs:
                    for sub in push_subs:
                        push_sent = _send_push_one(sub, push_title, push_body, "/", ld.get("phone","")) or push_sent

            # Telegram — always send (as backup)
            try:
                await app.bot.send_message(
                    chat_id=OWNER_CHAT_ID,
                    text=tg_text,
                    parse_mode="Markdown")
            except Exception as tg_err:
                logger.warning(f"Telegram send failed: {tg_err}")

            logger.info(f"Reminder sent: lead={ld.get('id')} push={push_sent}")
            db_update_lead(ld["id"], status=ST_PRE_REMINDED)
        except Exception as e:
            logger.error(f"pre_reminder [{ld.get('id')}]: {e}")
    for ld in db_get_leads_post_call():
        try:
            await app.bot.send_message(
                chat_id=OWNER_CHAT_ID,
                text=f"📞 *מה קרה עם {ld['name']}?*\n📱 {ld.get('phone', '—')}",
                reply_markup=outcome_kb(ld["id"]), parse_mode="Markdown")
            db_update_lead(ld["id"], status=ST_CALLED)
        except Exception as e:
            logger.error(f"post_call [{ld.get('id')}]: {e}")
    for ld in db_get_leads_followup_due():
        try:
            await app.bot.send_message(
                chat_id=OWNER_CHAT_ID,
                text=f"🔄 *זמן לחזור אל {ld['name']}!*\n📱 {ld.get('phone', '—')}",
                reply_markup=outcome_kb(ld["id"]), parse_mode="Markdown")
            db_update_lead(ld["id"], status=ST_CALLED)
        except Exception as e:
            logger.error(f"followup_due [{ld.get('id')}]: {e}")

def job_self_ping():
    """Ping own /health every 4 min to keep Render Free tier awake.
    Runs inside APScheduler so it works even when GitHub Actions is delayed."""
    if not APP_URL:
        return
    try:
        url = APP_URL.rstrip("/") + "/health"
        req = _ureq.Request(url, method="GET")
        resp = _ureq.urlopen(req, timeout=10)
        logger.info(f"Self-ping → {resp.status}")
    except Exception as e:
        logger.warning(f"Self-ping failed (non-critical): {e}")

async def job_morning_briefing(app):
    active = db_get_active()
    old    = db_get_old()
    if not active and not old:
        return
    now   = datetime.now(TZ)
    lines = [f"☀️ *בוקר טוב! {now.strftime('%d/%m/%Y')}*\n"]
    if active:
        lines.append(f"📋 *שיחות ותזכורות ({len(active)}):*")
        for ld in active[:10]:
            ct = str(ld.get("call_time") or ld.get("follow_up_time") or "")
            dt = _parse_dt(ct) if ct else None
            t  = f" ⏰{dt.strftime('%H:%M')}" if dt else ""
            lines.append(f"{STATUS_EMOJI.get(ld.get('status'), '❓')} *{ld['name']}*{t}")
    if old:
        lines.append(f"\n📁 *{len(old)} לידים ישנים*")
    await app.bot.send_message(chat_id=OWNER_CHAT_ID, text="\n".join(lines), parse_mode="Markdown")
    # Web Push briefing
    push_body = f"{len(active)} שיחות היום" if active else "אין שיחות היום"
    if old:
        push_body += f" · {len(old)} ישנים"
    push_to_all("☀️ בוקר טוב!", push_body)

async def job_weekly_report(app):
    s    = db_stats_since((datetime.now(TZ) - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S"))
    rate = f"{s['won_count']/s['total']*100:.0f}%" if s["total"] else "—"
    await app.bot.send_message(
        chat_id=OWNER_CHAT_ID,
        text=f"📊 *דוח שבועי*\n\n"
             f"לידים: {s['total']} | ✅ נסגרו: {s['won_count']} | ❌ אבדו: {s['lost']}\n"
             f"💰 הכנסות: ₪{s['won_amount']:,.0f} | המרה: {rate}\n\nשבוע מוצלח! 🍸",
        parse_mode="Markdown")

async def job_monthly_report(app):
    now  = datetime.now(TZ)
    s    = db_stats_since(now.replace(day=1, hour=0, minute=0, second=0).strftime("%Y-%m-%dT%H:%M:%S"))
    rate = f"{s['won_count']/s['total']*100:.0f}%" if s["total"] else "—"
    await app.bot.send_message(
        chat_id=OWNER_CHAT_ID,
        text=f"📆 *דוח חודשי — {now.strftime('%m/%Y')}*\n\n"
             f"לידים: {s['total']} | ✅ נסגרו: {s['won_count']} | ❌ אבדו: {s['lost']}\n"
             f"💰 הכנסות: ₪{s['won_amount']:,.0f} | המרה: {rate}\n\nחודש מוצלח! 💪",
        parse_mode="Markdown")

async def job_stale_leads(app):
    stale = db_get_stale_leads()
    if not stale:
        return
    lines = [f"🔔 *{len(stale)} לידים מחכים לעדכון (3+ ימים):*\n"]
    for ld in stale[:10]:
        try:
            upd      = datetime.strptime(ld["updated_at"][:19], "%Y-%m-%dT%H:%M:%S")
            days_ago = (datetime.now(TZ).replace(tzinfo=None) - upd).days
        except Exception:
            days_ago = "?"
        lines.append(f"📞 *{ld['name']}* — {ld.get('phone', '—')} _(לפני {days_ago} ימים)_")
    await app.bot.send_message(
        chat_id=OWNER_CHAT_ID,
        text="\n".join(lines) + "\n\nפתח את האפליקציה לעדכון סטטוס 👆",
        parse_mode="Markdown")
    # Web Push
    push_to_all("🔔 לידים ממתינים", f"{len(stale)} לידים ללא עדכון 3+ ימים")

async def job_backup(app):
    """Nightly off-database backup: send an Excel of ALL leads to the owner via Telegram."""
    try:
        leads = db_get_all()
        if not leads:
            return  # nothing to back up
        buf = build_excel(None)  # all users
        buf.seek(0)
        now = datetime.now(TZ)
        fname = f"backup_leads_{now.strftime('%Y-%m-%d')}.xlsx"
        await app.bot.send_document(
            chat_id=OWNER_CHAT_ID,
            document=buf,
            filename=fname,
            caption=f"💾 גיבוי יומי — {len(leads)} לידים ({now.strftime('%d/%m/%Y')})"
        )
        logger.info(f"Backup sent: {len(leads)} leads")
    except Exception as e:
        logger.error(f"Backup job failed: {e}")

def _build_bot_app() -> Application:
    """Build a configured Telegram Application with all handlers (no polling)."""
    menu_regex = (
        "^(📋 לידים ישנים"
        "|📊 סיכום"
        "|⏰ תזכורות פעילות"
        "|📅 היום"
        "|🔍 חיפוש ליד"
        "|📤 יצוא Excel"
        "|✏️ עריכת ליד)$"
    )
    conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^➕ הוספת ליד$"), add_start)],
        states={
            ASK_NAME:  [MessageHandler(filters.TEXT & ~filters.COMMAND, got_name)],
            ASK_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_phone)],
            ASK_TIME:  [MessageHandler(filters.TEXT & ~filters.COMMAND, got_time)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel_add),
            MessageHandler(filters.Regex(menu_regex), cancel_add),
        ],
        allow_reentry=True,
    )
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(conv)
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    return app


async def _setup_webhook() -> None:
    """Start the bot in webhook mode — eliminates Telegram Conflict on redeploy."""
    global _tg_app
    if not BOT_TOKEN or not OWNER_CHAT_ID:
        logger.warning("BOT_TOKEN/OWNER_CHAT_ID not set — Telegram bot disabled")
        return
    if not APP_URL:
        logger.warning("APP_URL not set — cannot register webhook; Telegram bot disabled")
        return
    try:
        _tg_app = _build_bot_app()
        await _tg_app.initialize()
        await _tg_app.start()
        webhook_url = f"{APP_URL.rstrip('/')}/telegram/webhook"
        await _tg_app.bot.set_webhook(
            url=webhook_url,
            allowed_updates=["message", "callback_query"],
            drop_pending_updates=True,
        )
        logger.info(f"Telegram webhook set → {webhook_url}")
    except Exception as e:
        logger.error(f"Telegram webhook setup failed: {e}", exc_info=True)
        _tg_app = None


async def _teardown_webhook() -> None:
    """Stop the bot (webhook URL is kept — next deploy's set_webhook will overwrite it)."""
    global _tg_app
    if _tg_app is None:
        return
    try:
        await _tg_app.stop()
        await _tg_app.shutdown()
        logger.info("Telegram bot stopped gracefully (webhook retained)")
    except Exception as e:
        logger.warning(f"Bot teardown error (non-critical): {e}")
    finally:
        _tg_app = None

# ─────────────────────────────────────────────────────────────────────────────
#  FASTAPI APPLICATION
# ─────────────────────────────────────────────────────────────────────────────

def job_cleanup_sessions():
    """Delete expired sessions nightly to prevent DB bloat."""
    try:
        with _db() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM sessions WHERE expires_at < NOW()")
            deleted = cur.rowcount
        logger.info(f"Session cleanup: removed {deleted} expired session(s)")
    except Exception as e:
        logger.error(f"Session cleanup failed: {e}")

@asynccontextmanager
async def lifespan(app: FastAPI):
    ensure_db()
    if WEBPUSH_AVAILABLE:
        get_vapid_keys()

    # Switch to webhook mode — no polling, no Conflict on redeploy
    await _setup_webhook()

    # Start scheduler in the uvicorn event loop (not a separate thread)
    scheduler = AsyncIOScheduler(timezone="Asia/Jerusalem")
    scheduler.add_job(job_self_ping,         "interval", minutes=4)
    scheduler.add_job(job_cleanup_sessions,  "cron",     hour=3, minute=0)
    if _tg_app:
        scheduler.add_job(job_check_reminders,  "interval", minutes=1,                  args=[_tg_app])
        scheduler.add_job(job_morning_briefing, "cron",     hour=9,  minute=0,          args=[_tg_app])
        scheduler.add_job(job_weekly_report,    "cron",     day_of_week="sun", hour=10, args=[_tg_app])
        scheduler.add_job(job_monthly_report,   "cron",     day=1,   hour=10,           args=[_tg_app])
        scheduler.add_job(job_stale_leads,      "cron",     hour=10, minute=30,         args=[_tg_app])
        scheduler.add_job(job_backup,           "cron",     hour=4,  minute=0,          args=[_tg_app])
    scheduler.start()
    logger.info("Scheduler ✓")

    logger.info("🍸 Bar Lead Manager v4.0 started")
    yield

    scheduler.shutdown(wait=False)
    await _teardown_webhook()

fastapi_app = FastAPI(title="Bar Lead Manager API v4", lifespan=lifespan)
_ALLOWED_ORIGINS = [
    "https://bar-lead-bot.onrender.com",
    "http://localhost:8080",
    "http://127.0.0.1:8080",
]
fastapi_app.add_middleware(
    CORSMiddleware,
    allow_origins=_ALLOWED_ORIGINS,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
    allow_credentials=True,
)

# ── Pydantic models ───────────────────────────────────────────────────────────

class LeadCreate(BaseModel):
    name:      str            = Field(..., min_length=1, max_length=200)
    phone:     str            = Field(..., min_length=1, max_length=50)
    call_time: Optional[str]  = None
    notes:     Optional[str]  = ""
    tags:      Optional[str]  = ""
    source:    Optional[str]  = ""
    package:   Optional[str]  = ""
    force:     Optional[bool] = False  # skip duplicate-phone check

    @field_validator("name", "phone", mode="before")
    @classmethod
    def strip_and_require(cls, v):
        if isinstance(v, str):
            v = v.strip()
        return v

class LeadUpdate(BaseModel):
    name:           Optional[str]   = Field(None, min_length=1, max_length=200)
    phone:          Optional[str]   = Field(None, min_length=1, max_length=50)
    call_time:      Optional[str]   = None
    status:         Optional[str]   = None
    sale_amount:    Optional[float] = None
    notes:          Optional[str]   = None
    follow_up_time: Optional[str]   = None
    tags:           Optional[str]   = None
    source:         Optional[str]   = None
    package:        Optional[str]   = None

class RegisterRequest(BaseModel):
    username:     str = Field(..., min_length=3, max_length=50)
    password:     str = Field(..., min_length=8)
    display_name: Optional[str] = ""

class LoginRequest(BaseModel):
    username: str
    password: str

class ForgotPasswordRequest(BaseModel):
    username: str

class ResetPasswordRequest(BaseModel):
    username:     str
    code:         str
    new_password: str = Field(..., min_length=8)

class PushSubscribeRequest(BaseModel):
    subscription: dict

def enrich(d: dict) -> dict:
    d["status_heb"]   = STATUS_HEB.get(d.get("status", ""), d.get("status", ""))
    d["status_emoji"] = STATUS_EMOJI.get(d.get("status", ""), "")
    d["source_heb"]   = SOURCE_HEB.get(d.get("source") or "", d.get("source") or "")
    d["package_heb"]  = PACKAGE_HEB.get(d.get("package") or "", d.get("package") or "")
    return d

# ── Admin dependency ──────────────────────────────────────────────────────────

_bearer = HTTPBearer(auto_error=False)

def get_current_user(credentials: Optional[HTTPAuthorizationCredentials] = Depends(_bearer)) -> dict:
    if not credentials:
        raise HTTPException(401, "נדרשת התחברות")
    user = db_get_user_by_token(credentials.credentials)
    if not user:
        raise HTTPException(401, "הסשן פג תוקף — התחבר מחדש")
    return user

def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if not user.get("is_admin"):
        raise HTTPException(403, "נדרשת הרשאת מנהל")
    return user

# ── Auth Routes ───────────────────────────────────────────────────────────────

@fastapi_app.get("/api/app-config")
def app_config():
    has_users = db_count_users() > 0
    return {
        "auth_type":         "login",
        "has_users":         has_users,
        "version":           "4.0",
        "open_registration": OPEN_REGISTRATION,
    }

# ── RATE LIMITER (in-memory, IP-based) ────────────────────────────────────
_rate_limit_store: dict[str, list[float]] = {}
_rate_limit_lock = threading.Lock()
_RATE_LIMIT_MAX     = 5    # max attempts
_RATE_LIMIT_WINDOW  = 900  # 15 minutes in seconds

def _check_rate_limit(request: Request) -> None:
    """Raise HTTP 429 if IP has exceeded login/register attempts."""
    ip = request.client.host if request.client else "unknown"
    now = time.time()
    with _rate_limit_lock:
        attempts = _rate_limit_store.get(ip, [])
        # Keep only attempts within the window
        attempts = [t for t in attempts if now - t < _RATE_LIMIT_WINDOW]
        if len(attempts) >= _RATE_LIMIT_MAX:
            wait = int(_RATE_LIMIT_WINDOW - (now - attempts[0]))
            raise HTTPException(
                status_code=429,
                detail=f"יותר מדי ניסיונות כניסה. נסה שוב בעוד {wait} שניות."
            )
        attempts.append(now)
        _rate_limit_store[ip] = attempts

def _clear_rate_limit(ip: str) -> None:
    """Clear attempts after successful auth."""
    with _rate_limit_lock:
        _rate_limit_store.pop(ip, None)

@fastapi_app.post("/api/auth/register", status_code=201)
def register(body: RegisterRequest, request: Request):
    _check_rate_limit(request)
    if not OPEN_REGISTRATION and db_count_users() > 0:
        raise HTTPException(403, "הרשמה סגורה — פנה למנהל המערכת")
    if db_get_user_by_username(body.username):
        raise HTTPException(400, "שם המשתמש כבר קיים")
    uid   = db_create_user(body.username.strip(), body.password, body.display_name or body.username)
    token = db_create_session(uid)
    user  = db_get_user_by_token(token)
    _clear_rate_limit(request.client.host if request.client else "unknown")
    return {"token": token, "user": user}

@fastapi_app.post("/api/auth/login")
def login(body: LoginRequest, request: Request):
    _check_rate_limit(request)
    user = db_get_user_by_username(body.username.strip())
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(401, "שם משתמש או סיסמה שגויים")
    _clear_rate_limit(request.client.host if request.client else "unknown")
    token = db_create_session(user["id"])
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "display_name": user["display_name"],
            "is_admin": user.get("is_admin", False),
        }
    }

@fastapi_app.post("/api/auth/logout")
def logout(credentials: Optional[HTTPAuthorizationCredentials] = Depends(_bearer)):
    if credentials:
        db_delete_session(credentials.credentials)
    return {"ok": True}

@fastapi_app.get("/api/auth/me")
def get_me(user: dict = Depends(get_current_user)):
    return user

@fastapi_app.post("/api/auth/forgot-password")
def forgot_password(body: ForgotPasswordRequest):
    user = db_get_user_by_username(body.username.strip())
    if not user:
        # Always return the same message to prevent username enumeration
        return {"message": "אם המשתמש קיים, קוד איפוס נשלח"}
    code = db_create_reset_token(user["id"])
    sent = False
    if SMTP_HOST:
        sent = send_email(
            user.get("email", "") or "",
            "קוד איפוס סיסמה — Bar Lead Manager",
            f"<p>קוד האיפוס שלך: <strong>{code}</strong></p>"
            f"<p>הקוד תקף לשעה אחת.</p>"
        )
    if not sent:
        # No SMTP configured — return code for dev/admin use only (log it server-side)
        logger.info(f"Password reset code for {body.username}: {code}")
        # Still return it so the user can reset (single-user / no email setup)
        return {"message": "קוד איפוס נוצר — בדוק את הלוגים של המנהל", "code": code}
    return {"message": "קוד איפוס נשלח לאימייל"}

@fastapi_app.post("/api/auth/reset-password")
def reset_password(body: ResetPasswordRequest):
    uid = db_verify_reset_code(body.username.strip(), body.code.strip())
    if not uid:
        raise HTTPException(400, "קוד לא תקין או שפג תוקפו")
    db_update_password(uid, body.new_password)
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM sessions WHERE user_id=%s", (uid,))
    return {"message": "סיסמה אופסה בהצלחה — התחבר מחדש"}

# ── Service Worker ─────────────────────────────────────────────────────────────

_SW_JS = r"""
self.addEventListener('install', () => self.skipWaiting());
self.addEventListener('activate', e => e.waitUntil(clients.claim()));

self.addEventListener('push', function(event) {
  var data = {};
  try { data = event.data.json(); } catch(e) {}
  var title = data.title || 'Bar Lead Manager 🍸';
  var opts = {
    body:      data.body || '',
    icon:      '/icon-192.png',
    badge:     '/icon-192.png',
    data:      { url: data.url || '/', phone: data.phone || '' },
    vibrate:   [300, 100, 300, 100, 300],
    tag:       'blm-reminder',
    renotify:  true,
    requireInteraction: true,
    silent:    false
  };
  if (data.phone) {
    opts.actions = [
      { action: 'call', title: '📞 התקשר' },
      { action: 'dismiss', title: 'סגור' }
    ];
  }
  event.waitUntil(self.registration.showNotification(title, opts));
});

self.addEventListener('notificationclick', function(event) {
  event.notification.close();
  var phone = event.notification.data && event.notification.data.phone;
  if (event.action === 'call' && phone) {
    event.waitUntil(clients.openWindow('tel:' + phone));
    return;
  }
  var url = (event.notification.data && event.notification.data.url) || '/';
  event.waitUntil(
    clients.matchAll({ type: 'window', includeUncontrolled: true }).then(function(list) {
      for (var i = 0; i < list.length; i++) {
        if ('focus' in list[i]) { list[i].focus(); return; }
      }
      if (clients.openWindow) return clients.openWindow(url);
    })
  );
});
"""

@fastapi_app.get("/sw.js")
def serve_sw():
    from fastapi.responses import Response
    return Response(
        content=_SW_JS,
        media_type="application/javascript",
        headers={"Service-Worker-Allowed": "/", "Cache-Control": "no-cache"}
    )

@fastapi_app.get("/manifest.json")
def serve_manifest():
    from fastapi.responses import FileResponse
    p = os.path.join(STATIC_DIR, "manifest.json")
    if os.path.exists(p):
        return FileResponse(p, media_type="application/manifest+json")
    raise HTTPException(404)

@fastapi_app.get("/icon-192.png")
def serve_icon_192():
    from fastapi.responses import FileResponse
    p = os.path.join(STATIC_DIR, "icon-192.png")
    if os.path.exists(p):
        return FileResponse(p, media_type="image/png", headers={"Cache-Control": "public,max-age=86400"})
    raise HTTPException(404)

@fastapi_app.get("/icon-512.png")
def serve_icon_512():
    from fastapi.responses import FileResponse
    p = os.path.join(STATIC_DIR, "icon-512.png")
    if os.path.exists(p):
        return FileResponse(p, media_type="image/png", headers={"Cache-Control": "public,max-age=86400"})
    raise HTTPException(404)

# ── Push notification routes ──────────────────────────────────────────────────

@fastapi_app.get("/api/push/vapid-key")
def get_vapid_public_key():
    _, pub = get_vapid_keys()
    return {"publicKey": pub}

@fastapi_app.post("/api/push/subscribe")
def push_subscribe(body: PushSubscribeRequest, user: dict = Depends(get_current_user)):
    db_save_push_subscription(user["id"], json.dumps(body.subscription))
    logger.info(f"Push subscription saved for user {user['id']}")
    return {"ok": True}

@fastapi_app.get("/api/push/status")
def push_status_api(user: dict = Depends(get_current_user)):
    subs = db_get_push_subscriptions_for_user(user["id"])
    return {"subscribed": len(subs) > 0, "webpush_available": WEBPUSH_AVAILABLE}

@fastapi_app.post("/api/push/test")
def push_test_api(user: dict = Depends(get_current_user)):
    if not WEBPUSH_AVAILABLE:
        raise HTTPException(503, "Web push not available on server")
    subs = db_get_push_subscriptions_for_user(user["id"])
    if not subs:
        raise HTTPException(400, "אין מנוי. לחץ 'הפעל התראות' קודם.")
    last_err = ""
    sent = 0
    for s in subs:
        try:
            _send_push_one(s, "✅ בדיקה הצליחה!", "מערכת ההתראות עובדת! 🎉", "/", "")
            sent += 1
        except Exception as e:
            last_err = str(e)[:300]
    if not sent:
        raise HTTPException(500, f"שגיאה: {last_err}" if last_err else "שליחה נכשלה")
    return {"ok": True}

@fastapi_app.get("/api/push/debug")
def push_debug(user: dict = Depends(get_current_user)):
    """Diagnostic endpoint — returns push system status for troubleshooting."""
    priv, pub = get_vapid_keys()
    subs = db_get_push_subscriptions_for_user(user["id"])
    all_subs = db_get_all_push_subscriptions()
    priv_format = "base64url (correct)" if priv and not priv.startswith("-----") else \
                  "PEM (converting)" if priv and priv.startswith("-----") else \
                  "not set"
    return {
        "webpush_available": WEBPUSH_AVAILABLE,
        "vapid_key_format": priv_format,
        "vapid_public_key_set": bool(pub),
        "vapid_public_key_prefix": pub[:20] + "..." if pub else None,
        "your_subscriptions": len(subs),
        "total_subscriptions_all_users": len(all_subs),
        "user_id": user["id"],
    }

# ── Lead Routes ───────────────────────────────────────────────────────────────

@fastapi_app.get("/health")
def health():
    return {"status": "ok"}

@fastapi_app.post("/telegram/webhook")
async def telegram_webhook(request: Request):
    """Receive Telegram updates via webhook (replaces polling — no Conflict on redeploy)."""
    if _tg_app is None:
        return {"ok": False}
    try:
        data = await request.json()
        update = Update.de_json(data, _tg_app.bot)
        await _tg_app.process_update(update)
    except Exception as e:
        logger.error(f"Telegram webhook error: {e}", exc_info=True)
    return {"ok": True}

@fastapi_app.get("/api/leads/today")
def get_today_api(user: dict = Depends(get_current_user)):
    return [enrich(r) for r in db_get_today(user["id"])]

@fastapi_app.get("/api/leads/calendar")
def get_calendar_leads(
    from_dt: Optional[str] = None,
    to_dt:   Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    now = datetime.now(TZ)
    if not from_dt:
        from_dt = now.strftime("%Y-%m-%dT00:00:00")
    if not to_dt:
        to_dt = (now + timedelta(days=30)).strftime("%Y-%m-%dT23:59:59")
    rows   = db_calendar_leads(user["id"], from_dt, to_dt)
    result = {}
    for r in [enrich(r) for r in rows]:
        dt = _parse_dt(str(r.get("call_time", "")))
        if dt:
            day = dt.strftime("%Y-%m-%d")
            if day not in result:
                result[day] = []
            result[day].append(r)
    return result

@fastapi_app.get("/api/leads")
def get_leads(
    status: Optional[str] = None,
    q:      Optional[str] = None,
    source: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    uid = user["id"]
    if q:
        rows = db_search(q, uid)
    elif status:
        statuses     = status.split(",")
        placeholders = ",".join(["%s"] * len(statuses))
        with _db() as conn:
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(
                "SELECT * FROM leads WHERE status IN (" + placeholders + ") AND user_id=%s ORDER BY call_time",
                statuses + [uid]
            )
            rows = [dict(r) for r in cur.fetchall()]
    else:
        with _db() as conn:
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            if source:
                cur.execute(
                    "SELECT * FROM leads WHERE user_id=%s AND source=%s ORDER BY call_time DESC",
                    (uid, source)
                )
            else:
                cur.execute(
                    "SELECT * FROM leads WHERE user_id=%s ORDER BY call_time DESC", (uid,)
                )
            rows = [dict(r) for r in cur.fetchall()]
    return [enrich(r) for r in rows]

@fastapi_app.get("/api/leads/{lid}")
def get_lead_api(lid: str, user: dict = Depends(get_current_user)):
    row = db_get_lead(lid)
    if not row:
        raise HTTPException(404, "Lead not found")
    return enrich(row)

@fastapi_app.post("/api/leads", status_code=201)
def create_lead(body: LeadCreate, user: dict = Depends(get_current_user)):
    created_by = user.get("display_name") or user.get("username", "")
    # Duplicate-phone check (skipped when force=True)
    if not body.force:
        digits = "".join(ch for ch in body.phone if ch.isdigit())
        if digits:
            with _db() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT name FROM leads WHERE user_id=%s "
                    "AND regexp_replace(phone, '\\D', '', 'g') = %s LIMIT 1",
                    (user["id"], digits)
                )
                row = cur.fetchone()
            if row:
                raise HTTPException(
                    status_code=409,
                    detail=f"ליד עם הטלפון הזה כבר קיים: {row[0]}"
                )
    lid = db_add_lead_full(
        body.name, body.phone, body.call_time,
        body.notes or "", body.tags or "", created_by,
        source=body.source or "", package=body.package or "",
        user_id=user["id"]
    )
    msg = "<b>ליד חדש!</b>\n" + body.name + "\n" + body.phone + "\nנוסף ע׳׳י " + created_by
    if body.call_time:
        msg += "\n" + body.call_time
    if body.source:
        msg += "\n" + SOURCE_HEB.get(body.source, body.source)
    if body.notes:
        msg += "\n" + body.notes
    _send_tg(msg)
    return {"id": lid, "message": "Lead created"}

@fastapi_app.post("/api/leads/import")
async def import_leads_csv(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("windows-1255", errors="replace")

    reader     = csv.DictReader(io.StringIO(text))
    imported   = 0
    errors     = []
    created_by = user.get("display_name") or user.get("username", "")

    for i, row in enumerate(reader, 1):
        try:
            name  = (row.get("name") or row.get("שם") or "").strip()
            phone = (row.get("phone") or row.get("טלפון") or "").strip()
            if not name or not phone:
                errors.append("row " + str(i+1) + ": missing name/phone")
                continue
            notes     = (row.get("notes") or row.get("הערות") or "").strip()
            tags      = (row.get("tags") or row.get("תגיות") or "").strip()
            source    = (row.get("source") or row.get("מקור") or "").strip()
            package   = (row.get("package") or row.get("חבילה") or "").strip()
            call_time = (row.get("call_time") or row.get("שעת שיחה") or "").strip() or None
            db_add_lead_full(
                name, phone, call_time, notes, tags, created_by,
                source=source, package=package, user_id=user["id"]
            )
            imported += 1
        except Exception as e:
            errors.append("row " + str(i+1) + ": " + str(e)[:60])

    return {"imported": imported, "errors": errors[:20]}

@fastapi_app.put("/api/leads/{lid}")
def update_lead_api(lid: str, body: LeadUpdate, user: dict = Depends(get_current_user)):
    existing = db_get_lead(lid)
    if not existing:
        raise HTTPException(404, "Lead not found")
    raw     = body.model_dump(exclude_unset=True)
    updates = {k: v for k, v in raw.items() if not (k in ("name", "phone") and not v)}
    if not updates:
        return {"message": "No changes"}
    if "call_time" in updates and updates["call_time"]:
        cur_status = existing.get("status", "")
        if cur_status in (ST_PRE_REMINDED, ST_CALLED) and "status" not in updates:
            updates["status"] = ST_NEW
    new_status = updates.get("status")
    if new_status in (ST_WON, ST_LOST, ST_OLD) and "follow_up_time" not in updates:
        updates["follow_up_time"] = None
    db_update_lead(lid, **updates)
    return {"message": "Updated"}

@fastapi_app.delete("/api/leads/{lid}")
def delete_lead_api(lid: str, user: dict = Depends(get_current_user)):
    db_delete_lead(lid)
    return {"message": "Deleted"}

# ── Stats ─────────────────────────────────────────────────────────────────────

@fastapi_app.get("/api/stats")
def get_stats(user: dict = Depends(get_current_user)):
    uid       = user["id"]
    now       = datetime.now(TZ)
    week_ago  = (now - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S")
    month_ago = now.replace(day=1, hour=0, minute=0, second=0).strftime("%Y-%m-%dT%H:%M:%S")

    ACTIVE_STATUSES = (ST_NEW, ST_PRE_REMINDED, ST_CALLED, ST_FOLLOW_UP)
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM leads WHERE user_id=%s", (uid,))
        total = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM leads WHERE user_id=%s AND status='won'", (uid,))
        won_total = cur.fetchone()[0]
        cur.execute("SELECT COALESCE(SUM(sale_amount),0) FROM leads WHERE user_id=%s AND status='won'", (uid,))
        rev_total = float(cur.fetchone()[0])
        # Average sale amount (won deals with a positive amount)
        cur.execute("SELECT COALESCE(AVG(sale_amount),0) FROM leads WHERE user_id=%s AND status='won' AND sale_amount>0", (uid,))
        avg_sale = float(cur.fetchone()[0])
        # Active pipeline count
        cur.execute(
            "SELECT COUNT(*) FROM leads WHERE user_id=%s AND status IN (%s,%s,%s,%s)",
            (uid, *ACTIVE_STATUSES)
        )
        active = cur.fetchone()[0]
        # Status breakdown
        cur.execute("SELECT status, COUNT(*) FROM leads WHERE user_id=%s GROUP BY status", (uid,))
        status_counts = {row[0]: row[1] for row in cur.fetchall()}
        # Average days from creation to close (won leads)
        cur.execute(
            "SELECT COALESCE(AVG(EXTRACT(EPOCH FROM (updated_at::timestamp - created_at::timestamp))/86400),0) "
            "FROM leads WHERE user_id=%s AND status='won'",
            (uid,)
        )
        avg_days = float(cur.fetchone()[0] or 0)
        # Last 7 days: leads created per day
        cur.execute(
            "SELECT to_char(created_at::timestamp,'YYYY-MM-DD') AS d, COUNT(*) "
            "FROM leads WHERE user_id=%s AND created_at::timestamp >= %s GROUP BY d",
            (uid, week_ago)
        )
        daily_map = {row[0]: row[1] for row in cur.fetchall()}

    # Build ordered 7-day series (oldest→newest)
    daily = []
    for i in range(6, -1, -1):
        d = now - timedelta(days=i)
        key = d.strftime("%Y-%m-%d")
        daily.append({"label": d.strftime("%d/%m"), "count": daily_map.get(key, 0)})

    weekly  = db_stats_since(week_ago,  uid)
    monthly = db_stats_since(month_ago, uid)

    with _db() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            "SELECT source, COUNT(*) AS cnt FROM leads WHERE user_id=%s AND source != '' GROUP BY source ORDER BY cnt DESC",
            (uid,)
        )
        source_rows = [dict(r) for r in cur.fetchall()]

    return {
        "total":              total,
        "won_total":          won_total,
        "rev_total":          rev_total,
        "avg_sale":           round(avg_sale),
        "active":             active,
        "avg_days_to_close":  round(avg_days, 1) if avg_days else None,
        "status_counts":      status_counts,
        "daily":              daily,
        "weekly":             weekly,
        "monthly":            monthly,
        "source_counts":      {r["source"]: r["cnt"] for r in source_rows},
        "source_conversion":  db_source_conversion(uid),
    }

@fastapi_app.get("/api/export/excel")
def export_excel(user: dict = Depends(get_current_user)):
    uid = user["id"] if not user.get("is_admin") else None
    buf = build_excel(uid)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leads.xlsx"}
    )

# ── Admin Routes ──────────────────────────────────────────────────────────────

@fastapi_app.get("/api/admin/users")
def admin_list_users(admin: dict = Depends(require_admin)):
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT u.id, u.username, u.display_name, u.role, u.created_at,
                   COUNT(l.id) as lead_count
            FROM users u
            LEFT JOIN leads l ON l.user_id = u.id
            GROUP BY u.id ORDER BY u.created_at DESC
        """)
        rows = [dict(zip([d[0] for d in cur.description], r)) for r in cur.fetchall()]
    return rows

@fastapi_app.get("/api/admin/stats")
def admin_stats(admin: dict = Depends(require_admin)):
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM leads")
        total_leads = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM users")
        total_users = cur.fetchone()[0]
        cur.execute("SELECT COALESCE(SUM(sale_amount),0) FROM leads WHERE status=\'won\'")
        total_revenue = float(cur.fetchone()[0])
    return {"total_leads": total_leads, "total_users": total_users, "total_revenue": total_revenue}

@fastapi_app.delete("/api/admin/users/{uid}")
def admin_delete_user(uid: str, admin: dict = Depends(require_admin)):
    with _db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM leads WHERE user_id=%s", (uid,))
        cur.execute("DELETE FROM users WHERE id=%s", (uid,))
    return {"message": "User deleted"}

@fastapi_app.get("/{full_path:path}")
async def serve_index(full_path: str):
    index_path = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return HTMLResponse("<h1>App loading...</h1>")

if __name__ == "__main__":
    uvicorn.run(fastapi_app, host="0.0.0.0", port=PORT)
